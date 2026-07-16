#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
İzin veri katmanı — Yıllık İzin girişleri için kaynak veriyi hazırlar/doğrular.
==============================================================================
Bu modül PORTALDAN BAĞIMSIZDIR (Playwright yok). Tek başına çalıştırılıp doğrulanabilir.
`izin_poc.py` (canlı Playwright betiği) bunu import eder.

KAYNAKLAR (2026-06-17 kullanıcı doğruladı):
  1. DETAY izin (TEK doğru kaynak): "Mayıs 2026 İzin Günleri Detayı v2.xlsx"
     sayfa "Yıllık İzin Excel": B=Ad-Soyad, C=Date(dd.mm.yyyy), D=Gün (1=tam, 0,5=yarım).
     487 satır / 190 kişi. Spesifik tarihli — portala BU girilir.
  2. META (lokasyon + Ar-Ge sınıfı): ana "05-...FİNAL_9SAAT.xlsx" → "Mayıs" sayfası
     A=AD SOYAD, F=Lokasyonu SGK, H=Ar-Ge/Destek/K.Dışı.

NEDEN İKİ DOSYA: detayda lokasyon/Ar-Ge YOK; hangi portala + sadece Ar-Ge filtresi ana sayfadan gelir.
LOKASYON = "Lokasyonu SGK" (DGS ile tutarlı — kişi her iki formda aynı portala girer). SGK↔TGB aynı
parkların farklı etiketi (Yıldız=YTP, TPIz=İYTE); TPI ikisinde de "TPI".

İSİM EŞLEŞTİRME: ana sayfa BÜYÜK HARF + farklı yazım → Türkçe-fold (ç/ğ/ı/ö/ş/ü→c/g/i/o/s/u, İ/I→i).
Evlilik/göbek-adı farkları fold ile de tutmaz → IZIN_ALIAS ile elle eşlenir (canlı tespit, 2026-06-17).

ÖZET "2026 Mayıs Yıllık İzinler 9Saat" sayfası GÜVENİLMEZ (3 TPI kişi eksik + None'lar) → KULLANILMAZ.
Verify-or-halt: kişinin tarih toplamı detay dosyanın kendi içinden tutarlı (dış özetle değil).
"""
from __future__ import annotations
import argparse
import datetime
import sys
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook

# Detay-dosya adı  ->  ana "Mayıs" sayfasındaki tam ad (BÜYÜK HARF). Fold ile tutmayan gerçek kişiler.
# Hepsi 2026-06-17'de fuzzy + lokasyon ile teyit edildi. (3'ü TPI: BAYRAKTAR, GÜLAY SEVER, ŞEVVAL ATAY)
IZIN_ALIAS: dict[str, str] = {
    "Hilal Sedef Özkanoğlu": "HİLAL SEDEF BAYRAKTAR",     # TPI — soyadı değişikliği
    "Gülay Göktaş Sever":    "GÜLAY SEVER",                # TPI — ara soyad
    "Şevval Yorulmaz Atay":  "HAYRİYE ŞEVVAL ATAY",        # TPI — göbek adı + ara soyad
    "Faruk Kurtuluş":        "SADETTİN FARUK KURTULUŞ",    # Yıldız — göbek adı
    "Merve Özen Dursun":     "MERVE DURSUN",               # ODTÜ — ara soyad
}
# Ana sayfada HİÇ bulunamayan (lokasyon/Ar-Ge teyidi YOK) — portal autocomplete'inde / kullanıcıyla teyit.
# Özet sayfa ikisini de TPI gösteriyor ama doğrulanmadı → güvenli tarafta: UNRESOLVED raporlanır, atlanmaz/zorlanmaz.
IZIN_UNKNOWN_META = {"Beyhannur Çeviral", "Zehra Tekin"}

_TR = {"ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i", "I": "i", "i": "i",
       "ö": "o", "Ö": "o", "ş": "s", "Ş": "s", "ü": "u", "Ü": "u",
       "â": "a", "Â": "a", "î": "i", "Î": "i", "û": "u", "Û": "u"}


def fold(s: str | None) -> str:
    """Türkçe-güvenli normalize: diakritik sil + küçült + boşluk daralt. ('ABDULBAKİ'≡'Abdülbaki')."""
    if not isinstance(s, str):
        return ""
    return " ".join("".join(_TR.get(c, c) for c in s.strip()).lower().split())


def _norm_date(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


@dataclass
class IzinPerson:
    detay_ad: str                                  # detay dosyadaki ad (portal autocomplete'i bununla aranır)
    portal_ad: str                                 # ana sayfadaki resmî ad (alias çözülmüş)
    lokasyon: str                                  # Lokasyonu SGK
    ar_ge: str
    gunler: list[tuple[str, float]] = field(default_factory=list)  # [(dd.mm.yyyy, 1.0|0.5), ...]

    @property
    def toplam_gun(self) -> float:
        return round(sum(g for _, g in self.gunler), 3)

    @property
    def tam_gunler(self) -> list[str]:
        return [d for d, g in self.gunler if g == 1]

    @property
    def yarim_gunler(self) -> list[str]:
        return [d for d, g in self.gunler if g == 0.5]


def read_izin_detail(path: str, sheet: str = "Yıllık İzin Excel") -> dict[str, list[tuple[str, float]]]:
    """Detay dosya → {ad: [(tarih, gün), ...]}. Başlık satırı 'Ad-Soyad' aranarak bulunur (sabit indeks değil)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # başlık satırı + kolon indeksleri
    hdr_i = None
    ci = {}
    for i, r in enumerate(rows):
        cells = {fold(c): j for j, c in enumerate(r) if isinstance(c, str)}
        if "ad-soyad" in cells and "date" in cells and "gun" in cells:
            hdr_i = i
            ci = {"ad": cells["ad-soyad"], "date": cells["date"], "gun": cells["gun"]}
            break
    if hdr_i is None:
        raise ValueError(f"{path}: 'Ad-Soyad/Date/Gün' başlık satırı bulunamadı.")
    out: dict[str, list[tuple[str, float]]] = defaultdict(list)
    bad: list = []
    for r in rows[hdr_i + 1:]:
        ad = r[ci["ad"]]
        if not ad or not isinstance(ad, str):
            continue
        d = _norm_date(r[ci["date"]])
        g = r[ci["gun"]]
        try:
            g = float(g)
        except (TypeError, ValueError):
            bad.append((ad, d, g)); continue
        if g not in (1.0, 0.5):
            bad.append((ad, d, g)); continue          # beklenmeyen gün değeri → atla + raporla
        out[ad.strip()].append((d, g))
    if bad:
        print(f"[İZİN] UYARI: {len(bad)} satır beklenmeyen gün değeri (1/0,5 dışı) → atlandı: {bad[:5]}",
              file=sys.stderr)
    return dict(out)


def load_person_meta(path: str, sheet: str = "Mayıs") -> dict[str, dict]:
    """Ana sayfa → {fold(ad): {ad, sgk, arge}}. A=ad, F=Lokasyonu SGK, H=Ar-Ge/Destek."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    meta: dict[str, dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ad = r[0]
        if not ad or not isinstance(ad, str):
            continue
        meta[fold(ad)] = {
            "ad": ad.strip(),
            "sgk": (str(r[5]).strip() if len(r) > 5 and r[5] else ""),
            "arge": (str(r[7]).strip() if len(r) > 7 and r[7] else ""),
        }
    return meta


def _is_arge(s: str) -> bool:
    return "arge" in (s or "").lower().replace("-", "").replace(" ", "")


def build_izin_targets(izin_path: str, meta_path: str, lokasyon: str = "TPI",
                       only_arge: bool = True) -> tuple[list[IzinPerson], list[dict]]:
    """
    Detay + meta birleştir, lokasyon (+Ar-Ge) filtrele.
    Dönüş: (hedef IzinPerson listesi [ada göre sıralı], unresolved [meta bulunamayan] listesi).
    """
    detail = read_izin_detail(izin_path)
    meta = load_person_meta(meta_path)
    alias_fold = {fold(k): v for k, v in IZIN_ALIAS.items()}

    targets: list[IzinPerson] = []
    unresolved: list[dict] = []
    for ad, gunler in detail.items():
        f = fold(ad)
        m = meta.get(f)
        if m is None and f in alias_fold:                # evlilik/göbek-adı → resmî ad üzerinden
            m = meta.get(fold(alias_fold[f]))
        if m is None:
            unresolved.append({"ad": ad, "neden": "ana sayfada eşleşmedi (alias gerekebilir)",
                               "gun": round(sum(g for _, g in gunler), 3)})
            continue
        if m["sgk"].upper() != lokasyon.upper():
            continue
        if only_arge and not _is_arge(m["arge"]):
            continue
        targets.append(IzinPerson(
            detay_ad=ad, portal_ad=m["ad"], lokasyon=m["sgk"], ar_ge=m["arge"],
            gunler=sorted(gunler, key=lambda x: x[0].split(".")[::-1])))
    targets.sort(key=lambda p: fold(p.detay_ad))
    return targets, unresolved


def _report(izin_path: str, meta_path: str, lokasyon: str):
    targets, unresolved = build_izin_targets(izin_path, meta_path, lokasyon)
    print(f"\n==== İZİN HEDEFLERİ — Lokasyon={lokasyon} (Ar-Ge) ====")
    print(f"Toplam: {len(targets)} kişi\n")
    tot_full = tot_half = 0
    for p in targets:
        tot_full += len(p.tam_gunler); tot_half += len(p.yarim_gunler)
        print(f"  {p.portal_ad:32s} | {p.toplam_gun:>4} gün "
              f"({len(p.tam_gunler)} tam, {len(p.yarim_gunler)} yarım) | {p.gunler}")
    print(f"\nTOPLAM: {tot_full} tam gün + {tot_half} yarım gün")
    # lokasyondan bağımsız: unresolved (meta yok) — TPI olabilecekler dahil
    flagged = [u for u in unresolved if u["ad"] in IZIN_UNKNOWN_META] + \
              [u for u in unresolved if u["ad"] not in IZIN_UNKNOWN_META]
    if flagged:
        print(f"\n⚠️  META BULUNAMAYAN {len(flagged)} kişi (lokasyon/Ar-Ge teyit edilemedi — portalda/elle kontrol):")
        for u in flagged:
            tag = " [özet TPI diyor]" if u["ad"] in IZIN_UNKNOWN_META else ""
            print(f"     - {u['ad']} ({u['gun']} gün){tag}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="İzin veri katmanı doğrulama raporu (portalsız)")
    ap.add_argument("--izin", default="Mayıs 2026 İzin Günleri Detayı v2.xlsx")
    ap.add_argument("--meta", default="05-TEKNOKENTLER - MAYIS 2026 - FİNAL_9SAAT (1).xlsx")
    ap.add_argument("--lokasyon", default="TPI")
    args = ap.parse_args()
    _report(args.izin, args.meta, args.lokasyon)

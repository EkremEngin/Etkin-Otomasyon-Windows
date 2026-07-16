#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
İzin veri katmanı v2 — TEK-DOSYA self-contained format (T.C.'li).
=================================================================
Firmadan gelen yeni standart izin Excel'ini okur. ESKİ (izin_data.py) iki-dosya + isim-eşleştirme
tasarımının "kişi atlama" bug'ını KÖKTEN çözer:

  ESKİ: izin-detay (isim+tarih)  +  ana-personel (lokasyon+Ar-Ge)  →  İSİMLE eşleştir.
        Evlilik/göbek adı fold ile tutmayınca kişi sessizce 'unresolved' → ATLANIRDI.
  YENİ: TEK dosya. Her satırda  B=T.C.  C=Ad-Soyad  D=Tarih  E=Gün(1|0,5)  F=TGB(park kodu).
        İsim eşleştirmesi YOK; kimlik T.C. ile KESİN; park F kolonunda yazılı → atlama kaynağı yok.

TASARIM İLKESİ — "yanlış/eksik veri yazmaktansa DUR":
  Dosya doğrulaması HER satırı denetler (T.C. checksum, park kodu, gün, tarih). Herhangi bir SERT
  hata varsa TÜM hataları toplar ve `DataError` fırlatır (portala HİÇ dokunmadan) → insan Excel'i
  düzeltir. Temiz dosya (Haziran 2026 gibi) sorunsuz geçer. Böylece "30'da 4-5 kişi kayboldu"
  imkânsız: ya hepsi geçerli & işlenir, ya da hata net raporlanıp durulur.

PORTAL-BAĞIMSIZLIK:
  - Ay/dönem Excel'deki tarihlerden OTOMATİK türetilir ("HAZİRAN 2026", ay_regex) — her ay kod değişmez.
  - Park kodu → portal/onay-gereksinimi PARKS registry'sinde. ARI/ODTÜ = Teknoera (excel upload) →
    otomasyon=False; raporlanır ama Playwright akışına sokulmaz.

Kullanım (portalsız doğrulama raporu):
  python3 izin_data_v2.py "~/Downloads/Haziran 2026 Yıllık İzinler.xlsx"
  python3 izin_data_v2.py <dosya> --park TPI        # tek park detayı
"""
from __future__ import annotations
import argparse
import datetime
import os
import sys
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Türkçe-güvenli normalize (izin_data.py ile aynı davranış)
# ---------------------------------------------------------------------------
_TR = {"ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i", "I": "i", "i": "i",
       "ö": "o", "Ö": "o", "ş": "s", "Ş": "s", "ü": "u", "Ü": "u",
       "â": "a", "Â": "a", "î": "i", "Î": "i", "û": "u", "Û": "u"}


def fold(s: str | None) -> str:
    if not isinstance(s, str):
        return ""
    return " ".join("".join(_TR.get(c, c) for c in s.strip()).lower().split())


# ---------------------------------------------------------------------------
# Park registry — Excel F kolonu (TGB) kodu → portal + onay gereksinimi
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Park:
    code: str          # Excel F kolonu kodu (TGB) — canonical
    ad: str            # uzun ad
    portal_url: str
    label: str         # done-dosyası eki + eski sistem etiketi (izin_done_<label>_<Ay>.txt)
    otomasyon: bool    # True = Playwright "Arge Portal" (SUKTORSOFT); False = Teknoera (excel upload)
    onay_pdf: str      # onaya-gönder dosya gereksinimi: "" (PDF'siz) | "per_person" (kişi-başı belge)
    evrak_tipi: str = ""   # PDF'li parkta yükleme "Evrak Tipi" radyo ETİKETİ (park-bazlı; value HARDCODE ETME)


PARKS: "OrderedDict[str, Park]" = OrderedDict([
    ("TPI",    Park("TPI",    "Teknopark İstanbul",          "https://argeportal.teknoparkistanbul.com.tr/", "TPI",    True,  "")),
    ("BV",     Park("BV",     "Bilişim Vadisi",              "https://argeportal.bilisimvadisi.com.tr/",     "BV",     True,  "")),
    ("İYTE",   Park("İYTE",   "Teknopark İzmir",             "https://argeportal.teknoparkizmir.com.tr/",    "TPIz",   True,  "per_person", "yıllık izin formu")),
    ("YTP",    Park("YTP",    "Davutpaşa Yıldız Teknopark",  "https://argeportal.yildizteknopark.com.tr/",   "Yıldız", True,  "per_person", "Ücretli Yıllık İzin Formu")),
    ("ULUTEK", Park("ULUTEK", "Bursa Ulutek Teknopark",      "https://argeportal.ulutek.com.tr/",            "Ulutek", True,  "per_person", "Yıllık İzin Dilekçesi")),
    ("ARI",    Park("ARI",    "ARI Teknokent",               "https://portal.ariteknokent.com.tr/",          "Arı",    False, "")),
    ("ODTÜ",   Park("ODTÜ",   "ODTÜ Teknokent",              "https://portal.odtuteknokent.com.tr/",         "ODTÜ",   False, "")),
])

# Eski etiket / farklı yazımları canonical TGB koduna çevir (--park argümanı esnek olsun)
_PARK_ALIAS = {
    "tpiz": "İYTE", "izmir": "İYTE", "iyte": "İYTE",
    "yildiz": "YTP", "ytp": "YTP", "davutpasa": "YTP",
    "ari": "ARI", "odtu": "ODTÜ",
    "tpi": "TPI", "bv": "BV", "bilisimvadisi": "BV", "ulutek": "ULUTEK",
}


def resolve_park(code: str | None) -> Park | None:
    """Excel kodu / kullanıcı argümanı → Park. Fold ile esnek eşleşir; bulunamazsa None."""
    if not code:
        return None
    f = fold(code)
    if f in _PARK_ALIAS:
        return PARKS[_PARK_ALIAS[f]]
    for p in PARKS.values():
        if fold(p.code) == f or fold(p.label) == f or fold(p.ad) == f:
            return p
    return None


_TURKCE_AY = {1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
              7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"}
# Done-dosyası eki için Title-case (str.capitalize() Türkçe-İ'yi bozar: "HAZİRAN"→"Hazi̇ran" birleşik nokta;
# eski konvansiyon "Mayıs" dotless-ı → explicit map ŞART, tutarlılık için).
_AY_KEY = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
           7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}


# ---------------------------------------------------------------------------
# T.C. Kimlik No doğrulama (11 hane + resmî checksum)
# ---------------------------------------------------------------------------
def tc_valid(tc) -> bool:
    s = str(tc).strip()
    if not (s.isdigit() and len(s) == 11):
        return False
    d = [int(x) for x in s]
    if d[0] == 0:
        return False
    if (sum(d[0:9:2]) * 7 - sum(d[1:8:2])) % 10 != d[9]:
        return False
    if sum(d[0:10]) % 10 != d[10]:
        return False
    return True


# ---------------------------------------------------------------------------
# Kişi modeli
# ---------------------------------------------------------------------------
@dataclass
class IzinPersonV2:
    tc: str                                              # 11 haneli T.C. (kimlik anahtarı)
    ad: str                                              # Ad-Soyad (portal autocomplete'i bununla/T.C. ile arar)
    park_code: str                                       # canonical TGB kodu (F kolonu)
    gunler: "list[tuple[str, float]]" = field(default_factory=list)  # [(dd.mm.yyyy, 1.0|0.5)]

    @property
    def portal_ad(self) -> str:                          # izin_poc.process_one_person uyumu
        return self.ad

    @property
    def toplam_gun(self) -> float:
        return round(sum(g for _, g in self.gunler), 3)

    @property
    def tam_gunler(self) -> "list[str]":
        return [d for d, g in self.gunler if g == 1]

    @property
    def yarim_gunler(self) -> "list[str]":
        return [d for d, g in self.gunler if g == 0.5]

    @property
    def park(self) -> Park:
        return PARKS[self.park_code]


class DataError(Exception):
    """Excel'de SERT veri hatası → portala dokunmadan DUR (tüm hatalar mesajda listelenir)."""


# ---------------------------------------------------------------------------
# Okuma + doğrulama
# ---------------------------------------------------------------------------
_HDR_KEYS = {"tc": ("t.c.", "tc", "tckn", "kimlik", "t.c"),
             "ad": ("ad-soyad", "ad soyad", "adsoyad", "isim"),
             "date": ("tarih", "date"),
             "gun": ("gun sayisi", "gun", "gün sayısı")}


def _match_hdr(cell: str) -> str | None:
    f = fold(cell)
    for key, variants in _HDR_KEYS.items():
        if f in (fold(v) for v in variants):
            return key
    return None


def _find_header(rows) -> tuple[int, dict]:
    """T.C. + Ad-Soyad + Tarih + Gün başlığını içeren satırı ve kolon indekslerini bul (sabit indeks değil)."""
    for i, r in enumerate(rows):
        ci = {}
        for j, c in enumerate(r):
            if isinstance(c, str):
                k = _match_hdr(c)
                if k and k not in ci:
                    ci[k] = j
        if {"tc", "ad", "date", "gun"} <= set(ci):
            # TGB/park kolonu: 'Gün'den sonraki ilk metin başlık ("TGB")
            for j, c in enumerate(r):
                if isinstance(c, str) and fold(c) in ("tgb", "park", "lokasyon", "tgb kodu"):
                    ci["park"] = j
                    break
            if "park" not in ci:                          # başlıkta yoksa gün'den bir sonraki kolonu varsay
                ci["park"] = ci["gun"] + 1
            return i, ci
    raise DataError("Başlık satırı bulunamadı ('T.C.' + 'Ad-Soyad' + 'Tarih' + 'Gün Sayısı' aranıyor). "
                    "Beklenen format: sütunlarda T.C., Ad-Soyad, Tarih, Gün Sayısı, TGB.")


def _pick_sheet(wb) -> str:
    """Başlığı içeren sayfayı bul (sheet adı 'Tabelle1' vb. değişebilir → içerikten seç)."""
    for sh in wb.sheetnames:
        rows = list(wb[sh].iter_rows(values_only=True, max_row=15))
        try:
            _find_header(rows)
            return sh
        except DataError:
            continue
    return wb.sheetnames[0]


def _norm_date(v):
    """Excel hücresi → (dd.mm.yyyy str, datetime|None). Tarih değilse (str, None)."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d.%m.%Y"), v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.strftime("%d.%m.%Y"), dt
        except ValueError:
            pass
    return s, None


def read_izin_v2(path: str, sheet: str | None = None, strict: bool = True):
    """
    Yeni tek-dosya izin formatını oku + doğrula + parka göre grupla.

    Dönüş: (by_park, meta)
      by_park : OrderedDict[park_code, list[IzinPersonV2]]  (ada göre sıralı)
      meta    : {ay, yil, donem_label, ay_regex, ay_key, toplam_kisi, toplam_satir, sheet}

    strict=True: SERT hata (geçersiz T.C. / bilinmeyen park / gün≠1,0.5 / tarih dönem-dışı) varsa
                 TÜMÜ toplanır ve DataError fırlatılır (portala dokunulmaz).
    """
    path = os.path.expanduser(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = sheet or _pick_sheet(wb)
    ws = wb[sh]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, ci = _find_header(rows)

    # 1) satırları çöz + ay/yıl histogramı (dönem tespiti)
    raw = []                                              # (rownum, tc, ad, date_str, dt, gun, park_code)
    month_hist = defaultdict(int)
    for n, r in enumerate(rows[hdr_i + 1:], start=hdr_i + 2):
        def cell(k):
            j = ci[k]
            return r[j] if j < len(r) else None
        ad = cell("ad")
        if not ad or not str(ad).strip():
            continue                                      # boş satır
        date_str, dt = _norm_date(cell("date"))
        gun = cell("gun")
        park_raw = cell("park")
        raw.append((n, cell("tc"), str(ad).strip(), date_str, dt, gun,
                    (str(park_raw).strip() if park_raw is not None else "")))
        if dt:
            month_hist[(dt.year, dt.month)] += 1

    if not raw:
        raise DataError(f"{os.path.basename(path)}: veri satırı yok (başlık satırı {hdr_i+1}).")

    # dönem = en sık ay
    (yil, ay), _ = max(month_hist.items(), key=lambda kv: kv[1]) if month_hist else ((None, None), 0)
    donem_label = f"{_TURKCE_AY.get(ay, '?')} {yil}" if ay else "?"
    ay_regex = rf"^\d{{2}}\.{ay:02d}\.{yil}$" if ay else r"^\d{2}\.\d{2}\.\d{4}$"

    # 2) satır doğrulama
    errors = []
    people: "OrderedDict[str, IzinPersonV2]" = OrderedDict()   # tc -> person
    name_by_tc = {}
    for (n, tc, ad, date_str, dt, gun, park_raw) in raw:
        tcs = str(tc).strip() if tc is not None else ""
        # T.C.
        if not tc_valid(tcs):
            errors.append(f"  satır {n}: geçersiz T.C. {tcs!r} (kişi: {ad}) — 11 hane + checksum tutmuyor")
            continue
        # kimlik tutarlılığı (aynı TC farklı isim yazımı → uyar ama durdurma; ilk yazımı kullan)
        if tcs in name_by_tc and fold(name_by_tc[tcs]) != fold(ad):
            errors.append(f"  satır {n}: T.C. {tcs} iki farklı isimle: {name_by_tc[tcs]!r} / {ad!r}")
        name_by_tc.setdefault(tcs, ad)
        # park
        park = resolve_park(park_raw)
        if park is None:
            errors.append(f"  satır {n}: bilinmeyen park kodu {park_raw!r} (kişi: {ad}, T.C. {tcs}). "
                          f"Geçerli: {', '.join(PARKS)}")
            continue
        # gün
        try:
            g = float(gun)
        except (TypeError, ValueError):
            errors.append(f"  satır {n}: gün değeri sayı değil: {gun!r} (kişi: {ad})")
            continue
        if g not in (1.0, 0.5):
            errors.append(f"  satır {n}: gün 1 veya 0,5 olmalı, gelen: {g} (kişi: {ad})")
            continue
        # tarih
        if dt is None:
            errors.append(f"  satır {n}: tarih çözümlenemedi: {date_str!r} (kişi: {ad})")
            continue
        if ay and (dt.year, dt.month) != (yil, ay):
            errors.append(f"  satır {n}: tarih {date_str} dönem ({donem_label}) dışında (kişi: {ad})")
            continue
        # birik
        p = people.get(tcs)
        if p is None:
            p = IzinPersonV2(tc=tcs, ad=ad, park_code=park.code)
            people[tcs] = p
        elif p.park_code != park.code:
            errors.append(f"  satır {n}: {ad} (T.C. {tcs}) iki farklı parkta: {p.park_code} / {park.code}")
        p.gunler.append((date_str, g))

    # aynı kişi-aynı tarih tekrarı (çift giriş) uyarısı
    for p in people.values():
        seen = defaultdict(list)
        for d, g in p.gunler:
            seen[d].append(g)
        for d, gs in seen.items():
            if len(gs) > 1:
                errors.append(f"  {p.ad} (T.C. {p.tc}): {d} tarihi {len(gs)} kez ({gs}) — çift satır?")

    if errors and strict:
        raise DataError(f"{os.path.basename(path)}: {len(errors)} veri hatası bulundu "
                        f"(portala DOKUNULMADI, önce Excel'i düzeltin):\n" + "\n".join(errors))

    # parka göre grupla, kişileri isimle sırala, günleri tarihe göre sırala
    by_park: "OrderedDict[str, list[IzinPersonV2]]" = OrderedDict((c, []) for c in PARKS)
    for p in people.values():
        p.gunler.sort(key=lambda x: x[0].split(".")[::-1])
        by_park[p.park_code].append(p)
    for c in by_park:
        by_park[c].sort(key=lambda x: fold(x.ad))
    by_park = OrderedDict((c, v) for c, v in by_park.items() if v)   # boş parkları at

    meta = {
        "ay": ay, "yil": yil, "donem_label": donem_label, "ay_regex": ay_regex,
        "ay_key": _AY_KEY.get(ay, "?") if ay else "?",   # done-dosyası eki (Haziran) — Türkçe-güvenli Title-case
        "toplam_kisi": len(people), "toplam_satir": len(raw), "sheet": sh,
        "errors": errors,
    }
    return by_park, meta


def build_targets(path: str, park: str | None = None, only_otomasyon: bool = True,
                  strict: bool = True):
    """Tek park (veya tümü) için hedef liste. park=None → tüm otomasyon parkları düz liste + by_park."""
    by_park, meta = read_izin_v2(path, strict=strict)
    if park:
        pk = resolve_park(park)
        if pk is None:
            raise DataError(f"Bilinmeyen park: {park!r}. Geçerli: {', '.join(PARKS)}")
        targets = by_park.get(pk.code, [])
        return targets, meta, pk
    return by_park, meta, None


# ---------------------------------------------------------------------------
# Portalsız doğrulama raporu
# ---------------------------------------------------------------------------
def _report(path: str, park: str | None):
    try:
        by_park, meta = read_izin_v2(path, strict=True)
    except DataError as e:
        print(f"\n❌ VERİ HATASI — durduruldu:\n{e}\n", file=sys.stderr)
        sys.exit(2)

    print("=" * 78)
    print(f"İZİN VERİ RAPORU — {os.path.basename(path)}")
    print(f"Sayfa: {meta['sheet']!r} | Dönem: {meta['donem_label']} (ay_regex={meta['ay_regex']})")
    print(f"Toplam: {meta['toplam_kisi']} kişi / {meta['toplam_satir']} izin-günü satırı")
    print("=" * 78)

    oto_kisi = tekno_kisi = 0
    for code, people in by_park.items():
        pk = PARKS[code]
        tag = "✅ OTOMASYON" if pk.otomasyon else "⏭️  Teknoera (excel upload — kapsam dışı)"
        pdf = f" [onay PDF: {pk.onay_pdf}]" if pk.otomasyon and pk.onay_pdf else ""
        satir = sum(len(p.gunler) for p in people)
        tam = sum(len(p.tam_gunler) for p in people)
        yarim = sum(len(p.yarim_gunler) for p in people)
        print(f"\n### {pk.code}  ({pk.ad})  —  {tag}{pdf}")
        print(f"    {len(people)} kişi | {satir} satır ({tam} tam + {yarim} yarım gün) | portal: {pk.portal_url}")
        if pk.otomasyon:
            oto_kisi += len(people)
        else:
            tekno_kisi += len(people)
        if park and resolve_park(park).code == code:
            for p in people:
                print(f"      - {p.ad:32s} T.C.{p.tc} | {p.toplam_gun:>4} gün | {p.gunler}")

    print("\n" + "=" * 78)
    print(f"OTOMASYON KAPSAMI: {oto_kisi} kişi  |  Teknoera (kapsam dışı): {tekno_kisi} kişi")
    print(f"Doğrulama: T.C.✓  gün(1/0,5)✓  tarih({meta['donem_label']})✓  park-kodu✓  → 0 SESSİZ ATLAMA")
    print("=" * 78)
    if not park:
        print("Tek park detayı için: --park TPI  (veya BV/İYTE/YTP/ULUTEK)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="İzin veri katmanı v2 — portalsız doğrulama raporu")
    ap.add_argument("excel", help="İzin Excel yolu (ör. ~/Downloads/Haziran 2026 Yıllık İzinler.xlsx)")
    ap.add_argument("--park", default=None, help="Tek park detayını dök (TPI/BV/İYTE/YTP/ULUTEK)")
    args = ap.parse_args()
    _report(args.excel, args.park)

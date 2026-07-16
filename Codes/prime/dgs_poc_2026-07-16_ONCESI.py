#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DGS (Dışarıda Geçirilen Süreler) Otomasyon — verify-or-halt sürümü
==================================================================
Teknopark ARGE Portalı'nda personelin "Dışarıda Geçirilen Süreler" girişini, firmanın
gönderdiği Excel puantajına göre otomatik dolduran betik.

TASARIM İLKESİ — "yanlış veri yazmaktansa DUR":
  Bu resmi bir vergi/teşvik beyanıdır. Her adım kendi sonucunu DOĞRULAR. Bir doğrulama
  başarısız olursa o kişinin formu KAYDEDİLMEZ (terk edilir) ve FAILED işaretlenir; sunucuya
  asla hatalı/eksik veri gitmez. Form "Kaydet (F3)" YALNIZCA tüm kontroller geçerse tıklanır.

ÇALIŞMA MODELİ (ATTENDED / gözetimli):
  - Cloudflare doğrulamasını + girişi SEN bir kez elle yaparsın. Betik Cloudflare'i ASLA geçmeye
    çalışmaz; görürse durur ve sana bırakır. Sayfa ASLA reload EDİLMEZ (reload Cloudflare'i tetikler).
  - Betik senin giriş yaptığın Chrome'a CDP ile bağlanır.
  - VARSAYILAN TASLAK seviyesinde durur. "Onaya Gönder" + e-imza HER ZAMAN sende kalır.

CANLI DOĞRULANMIŞ BULGULAR (2026-06-17, ABDULBAKİ+AHMET+ALPHAN canlı girildi):
  1. PROJE SEÇİMİ: serbest yazıp öneriye tıklamak gizli `Proje_Id`'yi SET ETMİYOR
     ("proje seçiniz" hatası + tikler tutmaz). DOĞRU: yaz → jQuery autocomplete('search') tetikle
     → çıkan <li>'ye tıkla → gizli `Proje_Id` != "0"/"" olduğunu DOĞRULA. (_autocomplete_pick)
  2. DIALOG SELECT'LERİ ÇİFT: BaslangicSaat/Dakika/BitisSaat/Dakika hem kartın sağ panelinde hem
     ✎ dialogunda var. Dialog'unkini KONUM ile ayırt et (ekran x < 500). Yanlış sete yazmak felaket.
  3. STALE DIALOG: ✎ açıp HEMEN okursan önceki günün verisini alırsın. dlgTarih == hedef olana
     kadar bekle; her gün teyit et.
  4. DIALOG KAYDET kimi zaman ilk tıklamada tutmaz → kaydedildi mi (dialog kapandı + satır toplam ==
     gerekli) DOĞRULA, tutmadıysa bir kez daha dene.
  5. ÇOKLU KİŞİ: form toolbar'ındaki "Yeni (F2)" personel autocomplete'ini BOZUYOR (sonraki kişide
     remote arama tetiklenmiyor, Çalışma Türü boşalıyor). Bu yüzden her kişi için menüden TEMİZ
     yeniden açıyoruz (reopen_fresh_form), form-Yeni KULLANMIYORUZ.
  6. Dönem 151 = MAYIS 2026. Görevlendirme Türü = Diğer. Çalışma Türü = 10766 Sy. CB Kararı.

Kurulum / çalıştırma için README.md'ye bak.
"""

from __future__ import annotations
import argparse
import os
import sys
import time
from dataclasses import dataclass

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout

# ----------------------------------------------------------------------------
# AYARLAR
# ----------------------------------------------------------------------------
CONFIG = {
    "cdp_url": os.environ.get("DGS_CDP", "http://localhost:9222"),  # paralel park için: DGS_CDP ile port ver (9222 BV / 9223 TPIz)
    "portal_url": "https://argeportal.teknoparkistanbul.com.tr/",

    "donem_text": None,               # main()'de --donem/prev_month_label() ile set edilir (park-bazlı Donem_Id
    "ay_regex": None,                 # değişir → etiketle seç, Plan §5.2). Artık hardcode YOK — her ay otomatik.

    "daily_target_min": 540,         # YEDEK varsayılan. Artık kişi-bazında program haritasından türetilir
                                     # (5 gün/45 saat → 540/9:00). Harita yoksa bu kullanılır.
    "include_saturday": False,       # kişi-bazında belirlenir; bu yalnızca yedek
    "assume_standard": False,        # Harita kişiyi KAÇIRIRSA (Personel Listesi okuması kararsız) standart
                                     # 5g/45s (540dk/9:00, cmt yok) VARSAY. GÜVENLİ: grid'in temiz-gün
                                     # varsayılanı (09:00) ile çapraz-doğrulanır (uymazsa durur). --assume-std ile aç.
    "include_destek": False,         # --include-destek: Ar-Ge olmayan (Destek) kişileri de işle (4691 teşvik).
    "allow_6day": False,             # 6 gün/45 saat (cumartesi + 7,5h) akışı CANLI TEST EDİLMEDİ → varsayılan
                                     # KAPALI: 6 günlük kişi görülürse DURUR (yanlış girmez). Test sonrası True yap.

    "prefer_start": "09:00",
    "slot_margin": int(os.environ.get("DGS_MARGIN", "2")),  # PDKS aralık sınırından boşluk (dk). DGS_MARGIN=0 ile PDKS-eksik düzeltme (boşluğu tam kapat).
    "pdks_buffer": int(os.environ.get("DGS_PDKS_BUFFER", "0")),  # PDKS gününe fazladan dk: SGK badge saniyesini aşağı yuvarlıyor (form 09:00=SGK 08:59) → +2 ile 9h'ı net geç (gün cap'liyor).

    "calisma_turu_text": "10766",    # Çalışma Türü autocomplete'inde aranacak (10766 Sy. CB Kararı)

    "dry_run": True,                 # True iken form "Kaydet (F3)" ASLA tıklanmaz.
    "auto_onay": False,              # --onayla: Kaydet'ten ~0.5sn sonra formdaki "Onaya Gönder"e bas (taslak+onay
                                     # TEK geçişte; buton zaten Kaydet'in dibinde). Onay gidemezse taslak DURUR.
}

LOG = "[DGS]"


def log(*a):
    print(LOG, *a, flush=True)


# ----------------------------------------------------------------------------
# DÖNEM (2026-07-14 eklendi): dgs_onaya.py/dgs_rapor_kontrol.py'deki KANITLANMIŞ desenin birebiri —
# hardcode "MAYIS 2026" yerine bugünden bir önceki takvim ayı otomatik türetilir; --donem ile override edilebilir.
# Excel formatı hep aynı kalacağı için (kullanıcı 2026-07-14) artık her ay tek kod değişikliği gerekmez.
# ----------------------------------------------------------------------------
TR_AYLAR = {1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
            7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"}
TR_AYLAR_TITLE = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
                  7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}
_TR_AYLAR_REV = {v: k for k, v in TR_AYLAR.items()}


def prev_month_label_parts(today=None) -> tuple[int, int]:
    """Bugünden bir önceki takvim ayı → (yıl, ay_no). dgs_onaya.prev_month_label ile aynı hesap."""
    import datetime
    today = today or datetime.date.today()
    y, m = today.year, today.month
    m -= 1
    if m == 0:
        m = 12
        y -= 1
    return y, m


def donem_label_and_regex(override: str | None) -> tuple[str, str, int]:
    """--donem verildiyse onu parse eder (ör. 'HAZİRAN 2026'), yoksa bir önceki ayı türetir.
    Döndürür: (donem_text, ay_regex, ay_no) — sheet adı varsayılanı için ay_no da lazım."""
    if override:
        parts = override.strip().upper().split()
        if len(parts) != 2 or parts[0] not in _TR_AYLAR_REV or not parts[1].isdigit():
            raise ValueError(f"--donem çözümlenemedi: {override!r} (örnek: 'HAZİRAN 2026')")
        ay_no, yil = _TR_AYLAR_REV[parts[0]], parts[1]
        label = f"{TR_AYLAR[ay_no]} {yil}"
    else:
        yil, ay_no = prev_month_label_parts()
        label = f"{TR_AYLAR[ay_no]} {yil}"
    return label, rf"^\d{{2}}\.{ay_no:02d}\.{yil}$", ay_no


class VerifyError(Exception):
    """Bir doğrulama (postcondition) başarısız oldu → kişi KAYDEDİLMEDEN atlanır."""


class CloudflareHalt(Exception):
    """Cloudflare/oturum gerekli → insan müdahalesi. Betik durur."""


# ----------------------------------------------------------------------------
# SAF MANTIK: zaman + çakışmayan slot algoritması (test edilebilir, canlı doğrulandı)
# ----------------------------------------------------------------------------
def to_min(hhmm: str) -> int:
    h, m = hhmm.strip().split(":")
    return int(h) * 60 + int(m)


def to_hhmm(mins: int) -> str:
    return f"{mins // 60:02d}:{mins % 60:02d}"


def free_gaps(busy: list[tuple[int, int]], day_start: int = 0, day_end: int = 24 * 60) -> list[tuple[int, int]]:
    busy = sorted(busy)
    gaps, cur = [], day_start
    for s, e in busy:
        if s > cur:
            gaps.append((cur, s))
        cur = max(cur, e)
    if cur < day_end:
        gaps.append((cur, day_end))
    return gaps


def place_outside(needed_min: int, busy: list[tuple[int, int]], prefer_start: str = "09:00",
                  margin: int = 2, day_start: int = 0, day_end: int = 24 * 60 - 1) -> tuple[int, int]:
    """
    needed_min'lik dışarıda penceresini, PDKS aralıklarıyla ÇAKIŞMAYAN ve sınırlarına DEĞMEYEN
    (>= margin boşluk) en erken uygun slota koyar. Tercihen prefer_start'tan sonra. Yer yoksa ValueError.
    """
    if needed_min <= 0:
        raise ValueError("needed_min <= 0")
    ps = to_min(prefer_start)
    busy_sorted = sorted(busy)
    busy_ends = {e for _, e in busy_sorted}
    busy_starts = {s for s, _ in busy_sorted}
    eff = []
    for gs, ge in free_gaps(busy_sorted, day_start, day_end):
        a = gs + (margin if gs in busy_ends else 0)
        b = ge - (margin if ge in busy_starts else 0)
        if b - a >= needed_min:
            eff.append((a, b))
    cands = [(max(a, ps), b) for a, b in eff if b - max(a, ps) >= needed_min]
    if cands:
        st = min(cands)[0]
        return st, st + needed_min
    if eff:
        # prefer'dan sonra yer yok (öğleden sonra kartı + after-exit 24:00'ı aşıyor gibi) → GİRİŞE BİTİŞİK:
        # en geç biten uygun gap'i seç, bloğu o gap'in SONUNA yasla (izole 00:00 bloğu yerine karta komşu).
        a, b = max(eff, key=lambda g: g[1])
        return b - needed_min, b
    raise ValueError(f"{to_hhmm(needed_min)} için çakışmayan slot yok. busy={busy}")


# ----------------------------------------------------------------------------
# EXCEL
# ----------------------------------------------------------------------------
@dataclass
class PersonRow:
    ad_soyad: str
    proje_adi: str
    eksik_puantaj: str
    ar_ge: str = ""
    lokasyon: str = ""
    bolum: str = ""
    tc: str = ""


def read_excel(path: str, sheet: str = "Mayıs") -> dict[str, PersonRow]:
    """A=AD SOYAD, B=TC Kimlik, E=BÖLÜM, F=Lokasyonu SGK, H=Ar-Ge/Destek, N=Eksik Puantaj, W=Proje Adı.
    TC (2026-07-14 eklendi): select_personel'de T.C.-öncelikli kimlik doğrulaması için (izin_poc'taki
    CANLI kanıtlanmış fix'in DGS karşılığı — isim evlilik/kızlık soyadı farkına dayanıklı seçim)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    out: dict[str, PersonRow] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ad = (row[0] or "").strip() if row[0] else ""
        if not ad:
            continue
        tc_val = row[1] if len(row) > 1 and row[1] is not None else ""
        out[ad.upper()] = PersonRow(
            ad_soyad=ad,
            bolum=(str(row[4]).strip() if len(row) > 4 and row[4] else ""),
            lokasyon=(str(row[5]).strip() if len(row) > 5 and row[5] else ""),
            ar_ge=(str(row[7]).strip() if len(row) > 7 and row[7] else ""),
            eksik_puantaj=(str(row[13]).strip() if len(row) > 13 and row[13] is not None else ""),
            proje_adi=(str(row[22]).strip() if len(row) > 22 and row[22] else ""),
            tc=(str(tc_val).strip() if tc_val != "" else ""),
        )
    log(f"Excel okundu: {len(out)} personel ({sheet})")
    return out


def is_ar_ge(p: PersonRow) -> bool:
    return "arge" in (p.ar_ge or "").lower().replace("-", "")


def person_schedule(name: str, sched_map: dict) -> tuple[int, bool]:
    """Kişinin (günlük_hedef_dk, cumartesi_dahil_mi)'sini döndürür. Standart dışıysa VerifyError (DUR).
      5 gün / 45 saat → (540, False)  [CANLI TEST EDİLDİ — 9:00/gün, hafta içi]
      6 gün / 45 saat → (450, True)   [allow_6day=True gerektirir; CANLI TEST EDİLMEDİ]
    Harita kişiyi içermiyorsa veya program standart dışıysa DURUR (yanlış varsayım yerine).
    NOT: günlük hedef ayrıca grid'in temiz-gün varsayılanıyla çapraz kontrol edilir (process_one_person)."""
    key = name.strip().upper()
    if key not in sched_map:
        # Diakritik/boşluk toleranslı eşleşme (İ↔I, Ç↔C, Ö↔O ...): ASCII-fold edip karşılaştır.
        # SADECE tek bir aday varsa kullan (belirsizse DUR — yanlış kişiyi seçme).
        import unicodedata
        def _fold(s):
            s = unicodedata.normalize("NFKD", s or "")
            s = "".join(c for c in s if not unicodedata.combining(c))
            return " ".join(s.upper().split())
        fk = _fold(name)
        cands = [k for k in sched_map if _fold(k) == fk]
        if len(cands) == 1:
            log(f"   (isim fold-eşleşti: '{name}' → portal '{cands[0]}')")
            key = cands[0]
        else:
            toks = set(fk.split())
            near = sorted({k for k in sched_map if toks & set(_fold(k).split())})
            hint = (" | yakın: " + "; ".join(near[:6])) if near else ""
            # Harita okuması KARARSIZ (kişileri kaçırabiliyor) ama FEV TR tamamen 5g/45s.
            # Tek aday yoksa (gerçekten yok) + --assume-std açıksa standart varsay; grid çapraz-doğrular.
            if len(cands) == 0 and CONFIG.get("assume_standard"):
                log(f"   (UYARI: '{name}' haritada yok → standart 5g/45s ({CONFIG['daily_target_min']}dk) VARSAYILDI; grid temiz-günüyle doğrulanacak.)")
                return CONFIG["daily_target_min"], False
            extra = " (BİRDEN FAZLA fold-eşleşme, belirsiz)" if len(cands) > 1 else ""
            raise VerifyError(f"{name}: Personel Listesi haritasında program bulunamadı (isim eşleşmedi){extra}.{hint} MANUEL.")
    gun, saat = sched_map[key]
    if gun <= 0 or (saat * 60) % gun != 0:
        raise VerifyError(f"{name}: {saat} saat / {gun} gün tam bölünmüyor — MANUEL.")
    daily = saat * 60 // gun
    if gun == 5 and saat == 45:
        return daily, False
    if gun == 6 and saat == 45:
        if not CONFIG["allow_6day"]:
            raise VerifyError(
                f"{name}: 6 gün/45 saat (cumartesi + 7,5h). Bu akış CANLI TEST EDİLMEDİ. "
                f"CONFIG['allow_6day']=True yapıp birlikte tek kişide test ettikten SONRA koş. Şimdilik MANUEL.")
        return daily, True
    raise VerifyError(f"{name}: {gun} gün / {saat} saat — standart dışı program. MANUEL gir.")


# ----------------------------------------------------------------------------
# PLAYWRIGHT — düşük seviye yardımcılar (canlı doğrulanmış JS yöntemleri)
# ----------------------------------------------------------------------------
def attach_browser(pw):
    log(f"Chrome'a bağlanılıyor: {CONFIG['cdp_url']}")
    browser = pw.chromium.connect_over_cdp(CONFIG["cdp_url"])
    ctx = browser.contexts[0] if browser.contexts else browser.new_context()
    # --- YEDEK (eski hali — "argeportal" TÜM parklarda geçtiği için ÇOK GEVŞEKTİ; birden çok portal
    #     sekmesi açıkken YANLIŞ sekmeye bağlanıyordu, ör. --park TPIz koşusu açık BV sekmesini sürüyor,
    #     TPIz kişileri BV listesinde bulunamıyordu → hepsi "autocomplete'te eşleşme yok"):
    # page = next((p for p in ctx.pages if "argeportal" in (p.url or "")), None)
    _want = CONFIG["portal_url"].split("//", 1)[-1].split("/", 1)[0].lower()   # ör. argeportal.teknoparkizmir.com.tr
    page = next((p for p in ctx.pages if _want in (p.url or "").lower()), None)
    if page is None:
        # Sekme yoksa aç — ama Cloudflare çıkarsa kullanıcı geçecek (betik geçmez).
        page = ctx.new_page()
        page.goto(CONFIG["portal_url"])
    page.bring_to_front()
    return browser, page


def assert_logged_in(page: Page):
    """Cloudflare/login ekranındaysak DURDUR. Betik ASLA Cloudflare geçmez, reload YAPMAZ."""
    title = (page.title() or "")
    body = ""
    try:
        body = page.inner_text("body")[:600]
    except Exception:
        pass
    if ("Güvenlik doğrulaması" in body or "Just a moment" in title or "Bir dakika" in title
            or "Gerçek kişi" in body or "GİRİŞ" in title.upper()):
        raise CloudflareHalt("Cloudflare/giriş ekranı. Elle geç, dashboard açıkken tekrar koş.")
    # dashboard teyidi (menüde PERSONEL olmalı) — Cloudflare sonrası render gecikebilir → kısa poll
    # (gerçek Cloudflare/login ise yukarıdaki body/title kontrolü zaten DURDURMUŞ olurdu)
    for _ in range(8):
        if page.locator("text=PERSONEL").count() > 0:
            return
        page.wait_for_timeout(1000)
    raise CloudflareHalt("Dashboard görülemedi (PERSONEL menüsü yok). Giriş yap, tekrar koş.")


def _open_menu_item(page: Page, item_text: str):
    """PERSONEL menüsünden bir öğeyi aç — Playwright GERÇEK click (CDP) ile. CANLI BULGU (2026-06-17):
    menü jQuery DELEGATED; JS .click() (synthetic) navigasyonu TETİKLEMİYOR, gerçek mouse click şart.
    Ayrıca dropdown öğeleri menü kapalıyken 'not visible' → önce PERSONEL'e gerçek-click (menü açılır),
    sonra GÖRÜNÜR öğeye (.filter(visible=True)) gerçek-click. (Doğrulandı: Personel Listesi açıldı, 51 satır.)"""
    last = None
    for _ in range(3):
        try:
            page.get_by_text("PERSONEL", exact=True).first.click(timeout=5000)
            page.wait_for_timeout(600)
            page.get_by_text(item_text, exact=True).filter(visible=True).first.click(timeout=5000)
            page.wait_for_timeout(1000)
            # menü AÇIK kalırsa #caixa içindeki <a>'lar form toolbar'ının üstüne biner ve GERÇEK click'i
            # yer ('Onaya Gönder' gitmez) → öğeye bastıktan sonra dropdown'ı kapat (2026-07-14 TPI bulgusu)
            close_nav_menu(page)
            return
        except Exception as e:
            last = e
            page.wait_for_timeout(500)
    raise VerifyError(f"Menü öğesi açılamadı: {item_text} ({last})")


def open_dgs_form(page: Page):
    """PERSONEL → Dışarıda Geçirilen Süreler Formu (AJAX menü; reload yok)."""
    # TEMİZ SLATE: bekleyen tüm pencereleri kapat (önceki run/dry-run/sil formu açık kalmış olabilir →
    # yenisi üstüne açılınca İKİ grid olur, querySelectorAll karışır = '42 gün'/yanlış sınıflandırma).
    for _ in range(8):
        n = page.evaluate(r"""()=>{const b=[...document.querySelectorAll('div.ui-dialog-titlebar')].filter(x=>x.offsetParent!==null);
              b.forEach(x=>{const c=x.querySelector('.fa-times,.ui-dialog-titlebar-close'); if(c)c.click();}); return b.length;}""")
        page.wait_for_timeout(300)
        if n == 0:
            break
    _open_menu_item(page, "Dışarıda Geçirilen Süreler Formu")
    page.wait_for_selector('select#Donem_Id, select[name="Donem_Id"]', timeout=30_000)
    wait_for_form_idle(page)   # menü/form açılışı AJAX yükleme → bitene kadar bekle (kullanıcı uyarısı 2026-06-18)
    page.wait_for_timeout(700)


def read_schedule_map(page: Page) -> dict:
    """PERSONEL → Personel Tanıtım Kartı → Personel Listesi: TÜM sayfaları gez,
    {AD SOYAD (upper): (haftalık_gün, haftalık_saat)} döndür. Bitince pencereyi kapat.
    Kolonlar CANLI doğrulandı (2026-06-17): Adı Soyadı=idx3, Haftalık Çalışma Süresi=idx12, Günü=idx13.
    GÜVENLİK AĞI: sayfalama eksik kalsa bile, haritada OLMAYAN kişi `person_schedule`'da DURDURUR
    (yanlış girilmez) — yani bu fonksiyonun kusuru veri bütünlüğünü bozmaz, sadece o kişiyi atlatır."""
    import json
    # Menü açma + liste yüklenmesi flaky olabiliyor → KOMPLE retry (aç → satır bekle; gelmezse pencereleri
    # kapat + tekrar aç). 3 deneme.
    ROWS_JS = r"""() => [...document.querySelectorAll('tr')].some(tr=>{
                if(tr.offsetParent===null || tr.children.length<14) return false;
                const c=[...tr.children];
                return /^\d+$/.test((c[12].textContent||'').trim()) && /^\d+$/.test((c[13].textContent||'').trim());})"""
    # TEMİZ SLATE (2026-06-18): önceki run/çöküşten kalan açık pencereler (Liste/Form) PERSONEL menüsünü
    # bloklayıp 'Menü öğesi açılamadı (timeout)' veriyor → menüyü açmadan önce tüm dialogları kapat.
    for _ in range(8):
        n = page.evaluate(r"""()=>{const b=[...document.querySelectorAll('div.ui-dialog-titlebar')].filter(x=>x.offsetParent!==null);
              b.forEach(x=>{const c=x.querySelector('.fa-times,.ui-dialog-titlebar-close'); if(c)c.click();}); return b.length;}""")
        page.wait_for_timeout(300)
        if n == 0:
            break
    loaded = False
    for attempt in range(3):
        _open_menu_item(page, "Personel Tanıtım Kartı")
        try:
            page.wait_for_function(ROWS_JS, timeout=10000)
            loaded = True
            break
        except PWTimeout:
            # açılmadı → açık Personel Listesi pencerelerini kapat ve tekrar dene
            page.evaluate(
                r"""() => {[...document.querySelectorAll('div.ui-dialog-titlebar')].filter(b=>b.offsetParent!==null
                      && /Personel Listesi/.test(b.textContent)).forEach(b=>{const x=b.querySelector('.fa-times,.ui-dialog-titlebar-close'); if(x)x.click();});}""")
            page.wait_for_timeout(600)
    if not loaded:
        raise VerifyError("Personel Listesi satırları yüklenmedi (program haritası okunamadı) — MANUEL kontrol.")
    page.wait_for_timeout(400)
    READ_JS = r"""() => {const out=[];
              [...document.querySelectorAll('tr')].forEach(tr=>{
                if(tr.offsetParent===null || tr.children.length<14) return;
                const c=[...tr.children].map(td=>td.textContent.replace(/\s+/g,' ').trim());
                const ad=c[3], saat=c[12], gun=c[13];
                if(ad && /^\d+$/.test(saat) && /^\d+$/.test(gun)) out.push([ad, +gun, +saat]);});
              return JSON.stringify(out);}"""
    sched, pages = {}, 0
    while pages < 40:
        rows = json.loads(page.evaluate(READ_JS))
        for ad, gun, saat in rows:
            sched[ad.upper()] = (gun, saat)
        pages += 1
        cur_first = rows[0][0] if rows else None
        moved = page.evaluate(
            r"""() => {const nx=[...document.querySelectorAll('a,button,span,i')].filter(e=>e.offsetParent!==null
                  && /(fa-(angle|chevron|caret|step)-right|fa-forward|next|ileri)/i.test((e.className||'')+' '+(e.title||''))
                  && !/disabled/i.test(e.className) && getComputedStyle(e).pointerEvents!=='none');
                if(!nx.length) return false; nx[nx.length-1].click(); return true;}""")
        if not moved:
            break
        # yeni sayfa yüklenene kadar bekle (ilk satır adı değişsin); değişmezse son sayfa
        try:
            page.wait_for_function(
                r"""(prev) => {const tr=[...document.querySelectorAll('tr')].find(t=>t.offsetParent!==null
                      && t.children.length>=14 && /^\d+$/.test((t.children[13].textContent||'').trim()));
                    return tr && tr.children[3].textContent.trim()!==prev;}""",
                arg=cur_first, timeout=4000)
        except PWTimeout:
            break
    page.evaluate(
        r"""() => {const bar=[...document.querySelectorAll('div.ui-dialog-titlebar')]
              .filter(b=>b.offsetParent!==null && /Personel Listesi/.test(b.textContent))[0];
            if(bar){const x=bar.querySelector('.fa-times, .ui-dialog-titlebar-close'); if(x)x.click();}}""")
    page.wait_for_timeout(400)
    log(f"Program haritası: {len(sched)} kişi ({pages} sayfa). 6-günlük: "
        f"{[k for k,(g,s) in sched.items() if g!=5] or 'yok'}")
    return sched


def enable_saturday(page: Page):
    """6 günlük kişi için 'Cumartesi Günlerini Dahil Et' kutusunu işaretle (grid'e cumartesileri ekler).
    CANLI TEST EDİLMEDİ — yalnız CONFIG['allow_6day']=True iken çağrılır."""
    page.evaluate(
        r"""() => {const cb=[...document.querySelectorAll('input[name="Cumartesi"]')].filter(e=>e.offsetParent!==null).pop();
            if(cb && !cb.checked){cb.click(); if(window.jQuery) jQuery(cb).trigger('change');}}""")
    page.wait_for_timeout(900)


def _eval_visible_set_donem(page: Page) -> str:
    """Görünür Donem_Id select'ini ETİKETE göre seç (park-bazlı Donem_Id değeri değişir → etiketle bul).
    Playwright select_option (ham `value=` bazı combobox'ta tutmuyor — §0.9) + value DOĞRULA. Park-bağımsız."""
    label = CONFIG["donem_text"]
    val = page.evaluate(
        r"""(lbl) => {
            const sel=[...document.querySelectorAll('#Donem_Id, select[name="Donem_Id"]')].filter(e=>e.offsetParent!==null).pop();
            if(!sel) return '__YOK__';
            const o=[...sel.options].find(o=>(o.text||'').trim()===lbl);
            return o ? o.value : '__BULUNAMADI__';
        }""", label)
    if val in ("__YOK__", "__BULUNAMADI__"):
        opts = page.evaluate(
            r"""()=>{const s=[...document.querySelectorAll('#Donem_Id, select[name="Donem_Id"]')].filter(e=>e.offsetParent!==null).pop();
                return s?[...s.options].map(o=>(o.text||'').trim()).filter(Boolean):[];}""")
        raise VerifyError(f"Dönem etiketi {label!r} bulunamadı (val={val}). Mevcut dönemler: {opts}")
    page.locator('select#Donem_Id:visible, select[name="Donem_Id"]:visible').last.select_option(value=val, timeout=6000)
    page.wait_for_timeout(400)
    got = page.evaluate(
        r"""()=>{const s=[...document.querySelectorAll('#Donem_Id, select[name="Donem_Id"]')].filter(e=>e.offsetParent!==null).pop();
            return s?s.value:'';}""")
    if got != val:
        raise VerifyError(f"Dönem set edilemedi (value={got!r} != {val!r}, label={label!r}).")
    return val


def list_set_donem_and_yeni(page: Page):
    """LIST ekranında Dönem (ETİKET 'MAYIS 2026', park-bağımsız) set + 'Yeni (F2)' → temiz kart açar.
    NOT: liste önce 'Dönem Seçmelisiniz' der; o yüzden Yeni'den ÖNCE dönem set edilir."""
    _eval_visible_set_donem(page)   # etikete göre seçer + içeride DOĞRULAR (uymazsa VerifyError)
    page.wait_for_timeout(300)
    clicked = page.evaluate(
        r"""() => {
            const b=[...document.querySelectorAll('a,button')].find(x=>x.offsetParent!==null
              && /^Yeni \(F2\)/.test(x.textContent.trim()) && /newButton/.test(x.className));
            if(!b) return false; b.click(); return true;
        }""")
    if not clicked:
        raise VerifyError("Liste 'Yeni (F2)' butonu bulunamadı.")
    page.wait_for_selector('input[name="string_Personel_Id"]', timeout=20_000)
    wait_for_form_idle(page)   # 'Yeni' kartı açılışı AJAX yükleme → bekle (kullanıcı uyarısı 2026-06-18)


def ensure_card_fresh(page: Page):
    """Kart temiz mi? (personel boş, grid boş, dönem doğru). Değilse VerifyError."""
    st = page.evaluate(
        r"""() => {
          const p=[...document.querySelectorAll('input[name="string_Personel_Id"]')].filter(e=>e.offsetParent!==null);
          const pv = p.length ? p[p.length-1].value.trim() : '__yok__';
          const don=[...document.querySelectorAll('#Donem_Id, select[name="Donem_Id"]')].filter(e=>e.offsetParent!==null).pop();
          const proje=document.querySelector('input[type=hidden][name="Proje_Id"]');
          return JSON.stringify({pv, donemText:don?(don.selectedOptions[0]?.text||'').trim():'?', projeId:proje?proje.value:'?'});
        }""")
    import json
    s = json.loads(st)
    if s["pv"] == "__yok__":
        raise VerifyError("Kart açılmadı (personel input yok).")
    if s["pv"] != "":
        raise VerifyError(f"Kart temiz değil (personel dolu: {s['pv']!r}).")
    if s["donemText"] != CONFIG["donem_text"]:
        # dönem kartta boş/yanlış gelebilir → etikete göre set et (park-bağımsız)
        _eval_visible_set_donem(page)
    return True


def _autocomplete_pick(page: Page, input_name: str, type_text: str, match_substr: str,
                       verify_hidden: str | None = None, label: str = "") -> str:
    """
    jQuery UI autocomplete'ten GÜVENİLİR seçim (canlı doğrulanmış yöntem):
      odakla → değeri yaz → autocomplete('search') tetikle → görünür <li> (match_substr) tıkla
      → (varsa) gizli alan (verify_hidden) != '0'/'' DOĞRULA.
    Dönüş: seçilen görünür değer. Başarısızsa VerifyError.
    """
    # 1) aktif (topmost = son) input'a odaklan + değeri yaz
    ok = page.evaluate(
        r"""({name, txt}) => {
          const el=[...document.querySelectorAll(`input[name="${name}"]`)].filter(e=>e.offsetParent!==null).pop();
          if(!el) return false;
          el.focus(); el.value=txt;
          if(window.jQuery){ try{ jQuery(el).autocomplete('search', txt); }catch(e){} }
          el.dispatchEvent(new Event('input',{bubbles:true}));
          el.dispatchEvent(new Event('keyup',{bubbles:true}));
          return true;
        }""", {"name": input_name, "txt": type_text})
    if not ok:
        raise VerifyError(f"{label}: input bulunamadı ({input_name}).")
    # 2) menü görünene kadar bekle (NOT: :visible jQuery pseudo'su raw querySelectorAll'da GEÇERSİZ;
    #    görünürlük offsetParent ile kontrol edilir)
    try:
        page.wait_for_function(
            r"""(sub) => {
              const lis=[...document.querySelectorAll('ul.ui-autocomplete li, ul.ui-menu li')].filter(li=>li.offsetParent!==null);
              return lis.some(li => li.textContent.toLowerCase().includes(sub.toLowerCase()));
            }""", arg=match_substr, timeout=8000)
    except PWTimeout:
        raise VerifyError(f"{label}: '{type_text}' için öneri çıkmadı (autocomplete sonuç yok).")
    # 3) eşleşen li'yi GERÇEK mouse event DİZİSİYLE seç. CANLI BULGU: jQuery UI autocomplete'in select
    #    callback'i (input değerini set + grid AJAX'ını tetikler) tek bir .click()'le güvenilir kalkmıyor;
    #    MCP'de mouseover→mousedown→mouseup→click dizisi çalıştı.
    picked = page.evaluate(
        r"""(sub) => {
          const li=[...document.querySelectorAll('ul.ui-autocomplete li, ul.ui-menu li')]
            .filter(x=>x.offsetParent!==null && x.textContent.toLowerCase().includes(sub.toLowerCase()))[0];
          if(!li) return false;
          const tgt=li.querySelector('a,div')||li;
          for(const t of ['mouseover','mouseenter','mousemove','mousedown','mouseup','click'])
            tgt.dispatchEvent(new MouseEvent(t,{bubbles:true,cancelable:true,view:window}));
          return true;
        }""", match_substr)
    if not picked:
        raise VerifyError(f"{label}: öneri '{match_substr}' tıklanamadı.")
    page.wait_for_timeout(1000)
    # 4) doğrula
    if verify_hidden:
        val = page.evaluate(
            r"""(n) => {const h=document.querySelector(`input[type=hidden][name="${n}"]`); return h?h.value:'';}""",
            verify_hidden)
        if not val or val == "0":
            raise VerifyError(f"{label}: seçim sonrası gizli '{verify_hidden}' set olmadı (={val!r}).")
        return val
    return "ok"


def ensure_gorevlendirme_diger(page: Page):
    """Görevlendirme Türü = Diğer olmalı. Değilse set et + doğrula."""
    res = page.evaluate(
        r"""() => {
          // radio'lar: DisGorevlendirmeTuru_Id (3=Diğer). Etikete göre de bulabiliriz.
          const radios=[...document.querySelectorAll('input[type=radio]')].filter(e=>e.offsetParent!==null);
          // "Diğer" etiketli olanı bul
          let diger=null;
          for(const r of radios){
            const lbl=(r.closest('label')||{}).textContent || (document.querySelector(`label[for="${r.id}"]`)||{}).textContent || '';
            if(/Diğer/i.test(lbl)) diger=r;
          }
          if(!diger) return 'radio-yok';
          if(!diger.checked){ diger.click(); }
          return diger.checked ? 'diger' : 'set-edilemedi';
        }""")
    if res not in ("diger",):
        raise VerifyError(f"Görevlendirme Türü 'Diğer' yapılamadı (={res}).")


def ensure_calisma_turu(page: Page):
    """Çalışma Türü = 10766 Sy. CB Kararı. Boşsa autocomplete ile set + doğrula."""
    cur = page.evaluate(
        r"""() => {const el=[...document.querySelectorAll('input[name="string_CalismaTuru_Id"]')].filter(e=>e.offsetParent!==null).pop(); return el?el.value:'__yok__';}""")
    if cur == "__yok__":
        return  # bu portalda alan yok say
    if "10766" in cur:
        return
    _autocomplete_pick(page, "string_CalismaTuru_Id", CONFIG["calisma_turu_text"],
                       "10766", label="Çalışma Türü")


_TR_FOLD = {'ç': 'c', 'Ç': 'c', 'ğ': 'g', 'Ğ': 'g', 'ı': 'i', 'İ': 'i', 'I': 'i',
            'ö': 'o', 'Ö': 'o', 'ş': 's', 'Ş': 's', 'ü': 'u', 'Ü': 'u'}
_PERS_FOLD_JS = (r"""const _fold=s=>{const tr={"""
                 + ",".join(f"'{k}':'{v}'" for k, v in _TR_FOLD.items())
                 + r"""};return s.replace(/̇/g,'').split('').map(c=>tr[c]||c).join('').toLowerCase().replace(/\s+/g,' ').trim();};""")


def _fold_tr(s: str) -> str:
    s = (s or "").replace("̇", "")
    return " ".join("".join(_TR_FOLD.get(c, c) for c in s).lower().split())


def _click_combo(page: Page) -> bool:
    """▼ 'Tümünü Göster' combo toggle'ına GERÇEK Playwright click. Klavye-autocomplete bazı isimlerde
    dropdown'ı TETİKLEMİYOR (izin_poc'ta 2026-07-09 CANLI kanıtlandı: BV/TPI'de 'öneri çıkmadı' straggler'ları
    fatigue değil, tam bu yüzdendi). DGS aynı jQuery-UI autocomplete widget'ını kullanıyor → aynı fallback."""
    try:
        inp = page.locator("input[name='string_Personel_Id']:visible").first
        inp.locator("xpath=following-sibling::a[contains(@class,'combo')]").first.click(timeout=3000)
        return True
    except Exception:
        try:
            page.locator("a.combo:visible").first.click(timeout=2000)
            return True
        except Exception:
            return False


_DGS_TARGET_VISIBLE_JS = (r"""(args) => {const [tcArg, toks] = args;
      const f2=tcArg?tcArg.slice(0,2):'', l2=tcArg?tcArg.slice(-2):''; %s
      return [...document.querySelectorAll('ul.ui-autocomplete li, ul.ui-menu li')].filter(x=>x.offsetParent!==null)
        .some(li=>{const raw=li.textContent||''; const m=raw.match(/(\d{2})\*+(\d{2})/);
          const tcM = tcArg && m && m[1]===f2 && m[2]===l2;
          const nm=_fold(raw.replace(/\s+[\d\*]{3,}.*$/,''));
          const nmM = toks.every(t=>nm.includes(t));
          return tcM || nmM;});}""" % _PERS_FOLD_JS)

_DGS_PICK_LI_BY_TC_JS = (r"""(args) => {const [tcArg, toks] = args;
      const f2=tcArg?tcArg.slice(0,2):'', l2=tcArg?tcArg.slice(-2):''; %s
      const lis=[...document.querySelectorAll('ul.ui-autocomplete li, ul.ui-menu li')].filter(x=>x.offsetParent!==null);
      const cand=lis.map(li=>{const raw=li.textContent||''; const m=raw.match(/(\d{2})\*+(\d{2})/);
        const nm=_fold(raw.replace(/\s+[\d\*]{3,}.*$/,''));
        return {li, text: raw.trim(), tcM:(tcArg && m && m[1]===f2 && m[2]===l2), ov:toks.filter(t=>nm.includes(t)).length};});
      let pick = tcArg ? cand.filter(c=>c.tcM) : cand.filter(c=>c.ov===toks.length);
      pick.sort((a,b)=>b.ov-a.ov);
      if(!pick.length) return JSON.stringify({ok:false, reason:(tcArg?'tc-maske eşleşmedi':'ad eşleşmedi')});
      const best=pick[0];
      if(tcArg && best.ov===0) return JSON.stringify({ok:false, reason:'T.C. eşleşti ama isim HİÇ tutmuyor (güvenlik dur)'});
      const tgt=best.li.querySelector('a,div')||best.li;
      for(const t of ['mouseover','mouseenter','mousemove','mousedown','mouseup','click'])
        tgt.dispatchEvent(new MouseEvent(t,{bubbles:true,cancelable:true,view:window}));
      return JSON.stringify({ok:true, picked:best.text, overlap:best.ov, viaTc:!!best.tcM});}""" % _PERS_FOLD_JS)


def select_personel(page: Page, ad_soyad: str, tc: str | None = None):
    """Personel seç — T.C.-ÖNCELİKLİ kimlik doğrulaması (2026-07-14 eklendi; izin_poc'ta 2026-07-09 CANLI
    kanıtlanan fix'in DGS karşılığı) + TAM-AD fold yedeği + grid teyidi.
    CANLI BULGU 2026-06-18: 3 kelimelik tam ad autocomplete'te BOŞ dönebiliyor; eski kod soyada düşünce
    'KAYA' → 'KIR-KAYAK' alt-dizesine eşleşip YANLIŞ kişiyi (EMİN→ECE) seçmişti.
    CANLI BULGU 2026-07-09 (izin, aynı kök-neden DGS'e de geçerli): isim EVLİLİK/KIZLIK soyadı farkı yüzünden
    ('GAMZE AYTEKİN' portalda, Excel'de 'Gamze Aytekin Makam') isim-arama yanlış kişiyi bulur/hiç bulamaz →
    'öneri çıkmadı' denip SESSİZCE atlanır (TPI Haziran turunda 6 kişi). T.C. (Excel kolon B) varsa dropdown'daki
    MASKELİ T.C. ('25*******28') ile ilk2+son2 eşleştirilir — isim farkına DAYANIKLI. T.C. yoksa/eşleşmezse
    eski TAM-AD fold davranışına düşer (davranış SADECE güçlenir, geriye dönük uyumlu).
    Ayrıca: birden çok sorgu adayı dene (tam ad, ilk-iki, ilk kelime, soyad) + klavye-arama dropdown'ı
    tetiklemezse ▼ combo fallback + seçimden sonra kimliği DOĞRULA; tutmazsa DUR."""
    intended = _fold_tr(ad_soyad)
    tokens = intended.split()
    tcs = str(tc).strip() if tc else ""
    use_tc = tcs.isdigit() and len(tcs) == 11
    tc_arg = tcs if use_tc else ""
    parts = ad_soyad.split()
    cands = [ad_soyad]
    if len(parts) >= 3:
        cands.append(" ".join(parts[:2]))
    cands.append(parts[0])
    if len(parts) > 1 and parts[-1] not in cands:
        cands.append(parts[-1])
    import json
    picked = False
    last_err = ""
    for q in cands:
        page.evaluate(
            r"""({name, txt}) => {const el=[...document.querySelectorAll(`input[name="${name}"]`)].filter(e=>e.offsetParent!==null).pop();
                if(el){el.focus(); el.value=txt; if(window.jQuery){try{jQuery(el).autocomplete('search',txt);}catch(e){}}
                  el.dispatchEvent(new Event('input',{bubbles:true})); el.dispatchEvent(new Event('keyup',{bubbles:true}));}}""",
            {"name": "string_Personel_Id", "txt": q})
        page.wait_for_timeout(1300)
        li_ok = page.evaluate(_DGS_TARGET_VISIBLE_JS, [tc_arg, tokens])
        if not li_ok and _click_combo(page):
            page.wait_for_timeout(800)
            li_ok = page.evaluate(_DGS_TARGET_VISIBLE_JS, [tc_arg, tokens])
        if not li_ok:
            last_err = f"'{q}' → dropdown'da hedef görünmedi"
            continue
        pick = json.loads(page.evaluate(_DGS_PICK_LI_BY_TC_JS, [tc_arg, tokens]))
        if not pick.get("ok"):
            last_err = f"'{q}' → {pick.get('reason')}"
            continue
        if use_tc and pick.get("viaTc") and pick.get("overlap", 0) < len(tokens):
            log(f"  (T.C.-eşleşme: portal '{pick.get('picked')}' ↔ Excel '{ad_soyad}' — isim tam tutmuyor "
                f"(evlilik/kızlık soyadı olabilir), T.C. ile teyitli)")
        picked = True
        break
    if not picked:
        raise VerifyError(f"Personel autocomplete'te eşleşme bulunamadı: {ad_soyad}"
                          + (f" (T.C. ...{tcs[-4:]})" if use_tc else "") + f" — {last_err} — MANUEL.")
    page.wait_for_timeout(1000)
    # KİMLİK DOĞRULAMA: seçilen personel alanı intended ile TAM eşleşmeli YA DA (T.C. ile seçildiyse) en az
    # bir isim-token örtüşmeli (ÇİFT-KONTROL — JS pick zaten overlap==0'ı reddetti, burada ikinci katman).
    selval = page.evaluate(
        r"""()=>{const e=[...document.querySelectorAll('input[name="string_Personel_Id"]')].filter(x=>x.offsetParent!==null).pop(); return e?e.value:'';}""")
    name_exact = intended in _fold_tr(selval)
    if not name_exact:
        overlap = sum(1 for t in tokens if t in _fold_tr(selval))
        if not (use_tc and overlap > 0):
            raise VerifyError(f"Personel KİMLİK doğrulaması BAŞARISIZ: seçilen={selval!r} beklenen={ad_soyad!r} — MANUEL.")
        log(f"  (isim T.C. ile teyitli ama tam tutmuyor: seçilen={selval!r} beklenen={ad_soyad!r} — GÜVENLİ)")
    # grid (gün satırları) AJAX ile dolar — görünene kadar bekle (sabit timeout yerine)
    try:
        page.wait_for_function(
            r"""(re)=>[...document.querySelectorAll('tr')].some(tr=>tr.children[2]
                  && new RegExp(re).test((tr.children[2].textContent||'').trim()))""",
            arg=CONFIG["ay_regex"], timeout=12000)
    except PWTimeout:
        raise VerifyError(f"Personel seçildi ama grid boş (gün satırı gelmedi): {ad_soyad}")
    wait_for_form_idle(page)
    n = len(read_grid_rows(page))
    log(f"  Personel seçildi+doğrulandı: {ad_soyad} (grid={n} gün)")


def set_proje(page: Page, proje_adi: str):
    """Proje seç + gizli Proje_Id'nin set olduğunu DOĞRULA. (En kritik adım!)"""
    if not proje_adi:
        raise VerifyError("Excel'de proje adı boş.")
    # benzersiz, ayırt edici bir substring seç (ortadan ~40 karakter yeterli; 'contains' filtresi)
    needle = proje_adi.strip()
    typed = needle[:40]
    pid = _autocomplete_pick(page, "string_Proje_Id", typed, needle[:25],
                             verify_hidden="Proje_Id", label="Proje")
    wait_for_form_idle(page, settle=300)   # proje seçimi ~2.5sn grid reload tetikler — bitmeden tiklersen tikler silinir (KRİTİK: settle yüksek tut)
    log(f"  Proje seçildi (Proje_Id={pid})")


# ----------------------------------------------------------------------------
# GRID OKUMA (sabit kolon indeksleri — canlı doğrulandı)
#   td[2]=Tarih, td[3]=Başlangıç, td[4]=Bitiş, td[9]=Toplam(dışarıda), td[17]=Kesinleşmiş(içeride)
# ----------------------------------------------------------------------------
def read_grid_rows(page: Page) -> list[dict]:
    import json
    s = page.evaluate(
        r"""(ayRegex) => {
          const re=new RegExp(ayRegex);
          const rows=[];
          document.querySelectorAll('tr').forEach(tr=>{const td=tr.children; if(td.length<18)return;
            const t=(td[2]?.textContent||'').trim(); if(!re.test(t))return;
            const cb=tr.querySelector('input[type=checkbox][name^="module_update_"]');
            rows.push({tarih:t, toplam:(td[9]?.textContent||'').trim(),
              aciklama:(td[12]?.textContent||'').trim(),
              kesin:(td[17]?.textContent||'').trim(),
              checked:cb?cb.checked:null, disabled:cb?cb.disabled:true});});
          return JSON.stringify(rows);
        }""", CONFIG["ay_regex"])
    return json.loads(s)


def wait_for_form_idle(page: Page, timeout: int = 20000, settle: int = 80):
    """Yükleme bitene kadar bekle — pop-up/dialog/grid/save açılışında 2-3 sn 'yükleme ekranı' dönebiliyor.
    KULLANICI UYARISI (2026-06-18): script aceleci davranıp loading bitmeden müdahale edince 'tutmadı/takıldı'
    sanıyoruz (Chrome yorgun DEĞİL). O yüzden HER pop-up/işlem sınırında burayı çağır.
    GÖSTERGE: görünür .form_loading_panel/.pReload.loading kaybolması (portalın gerçek yüklenme overlay'i —
    form aç, grid reload, dialog aç/kaydet hepsinde çıkar).
    🔴 KRİTİK BULGU (2026-06-18 canlı): jQuery.active bu portalda KALICI bağlantılar (SignalR/poll) yüzünden
    HİÇ 0 olmuyor (ölçüm: sabit 5). Eski 'jQuery.active==0' koşulu bu yüzden HER çağrıda 20s timeout yiyordu
    → kişi başı 8-9 DK! O koşul KALDIRILDI. Panel + downstream doğrulamalar (poll, İstenilen==hesap, satır toplamı)
    yeterli güvence. Bu portala özgü: jQuery.active'i idle göstergesi olarak ASLA kullanma.
    KRİTİK ESKİ BULGU: proje seçilince ~2.5 sn grid reload → o sırada tik atarsan tikler silinir → İstenilen 0:00."""
    try:
        page.wait_for_function(
            r"""() => ![...document.querySelectorAll('.form_loading_panel, .pReload.loading')]
                  .some(e=>e.offsetParent!==null)""",
            timeout=timeout)
    except PWTimeout:
        pass  # yine de devam; downstream doğrulamalar (İstenilen==hesap, poll, satır toplamı) yanlışı yakalar
    page.wait_for_timeout(settle)   # panel kaybolduktan sonra küçük DOM-render tamponu (proje-reload çağrısı settle'ı yükseltir)


def tick_clean_days(page: Page) -> int:
    """SADECE TEMİZ günleri tikle: kesin (PDKS, td[17]) değeri OLMAYAN, aktif, tiksiz satırlar.
    KRİTİK BULGU (2026-06-17): bazı PDKS günleri (örn. içeride çok düşükse) checkbox'ı ENABLED gelir;
    'disabled değilse tikle' demek o PDKS gününü yanlışça 9h clean gibi tikler. Ayrım `kesin` ile yapılır.
    (Proje seçili OLMALI — yoksa tikler İstenilen'e saymaz.)"""
    n = page.evaluate(
        r"""(re) => {
          const rx=new RegExp(re); let count=0;
          document.querySelectorAll('tr').forEach(tr=>{const td=tr.children; if(td.length<18) return;
            const t=(td[2]?.textContent||'').trim(); if(!rx.test(t)) return;
            const kesin=(td[17]?.textContent||'').trim();           // PDKS/izin/yarım-tatil → atla
            const ac=(td[12]?.textContent||'').trim();              // açıklamalı gün (Pdks/İzin/tatil) → atla
            const cb=tr.querySelector('input[type=checkbox][name^="module_update_"]');
            if(cb && !cb.disabled && !cb.checked && cb.offsetParent!==null && !kesin && !ac){ cb.click(); count++; }});
          return count;
        }""", CONFIG["ay_regex"])
    # tuttu mu? (sadece temiz günlerin sayısı kadar tikli olmalı)
    after = read_grid_rows(page)
    clean_checked = [r for r in after if r["checked"] and not r["kesin"]]
    if len(clean_checked) < n:
        raise VerifyError(f"Temiz gün tikleri tutmadı (tiklenen={n}, tikli temiz={len(clean_checked)}). Proje seçili mi?")
    return n


def _open_edit(page: Page, tarih: str) -> bool:
    ok = page.evaluate(
        r"""(d) => {const r=[...document.querySelectorAll('tr')].find(tr=>tr.children[2]&&tr.children[2].textContent.trim()===d);
              if(!r) return false; const ed=r.querySelector('span.edit-row'); if(!ed) return false; ed.click(); return true;}""",
        tarih)
    if ok:
        wait_for_form_idle(page)   # ✎ dialog ~2-3 sn yükleme ekranı dönebilir → bekle (kullanıcı uyarısı 2026-06-18)
    return ok


def _read_dialog(page: Page, tarih: str, expected_inside_min: int):
    """STALE guard (KRİTİK — canlı 2026-06-17/18 bulgusu): ✎ açınca tarih input'u Kesinleşmiş tablodan
    ÖNCE güncelleniyor → yalnız tarih+toplam'a bakmak STALE veriyi geçirebiliyor. FELAKET ÖRNEĞİ (ABDULBAKİ
    21↔22.05): iki günün içeride TOPLAMI tesadüfen eşit (8:45) ama aralıkları farklı → 22'yi açınca 21'in
    aralıkları okundu, sum==kesin geçti, yanlış pencere → portal 'pdks kaydı mevcuttur' reddetti.
    ÇÖZÜM: ham 'Giriş Çıkış' DAMGA satırları TARİH taşır ('22.05.2026 08:03 G PDKS') → dialog tarihi==tarih
    VE tüm PDKS damgalarının tarihi==tarih VE aralık toplamı==kesin OLANA KADAR bekle.
    Dönüş: (intervals[min,min]) veya VerifyError."""
    import json
    # NOT (2026-06-18): kesinleşmiş aralık TÜRLERİ: PDKS (badge), Ücretli (izin) vb. → regex hepsini almalı
    # (3×HH:MM + harf-tür). PDKS DAMGALARI tarih taşır (STALE kontrolü); izin gününde damga OLMAYABİLİR.
    js = r"""(d) => {
        const xs=[...document.querySelectorAll('input')].filter(i=>i.offsetParent!==null && /^\d{2}\.\d{2}\.\d{4}$/.test((i.value||'').trim()));
        if(xs.length!==1 || xs[0].value.trim()!==d.tarih) return false;
        // PDKS damgaları VARSA hepsi hedef tarihli olmalı (izin gününde hiç damga olmayabilir → izinli OK)
        let foreign=false;
        document.querySelectorAll('tr').forEach(tr=>{const t=tr.innerText.replace(/\s+/g,' ').trim();
          const m=t.match(/^(\d{2}\.\d{2}\.\d{4})\s+\d{2}:\d{2}\s+\S+\s+(PDKS|Manuel|Elle)/);
          if(m && m[1]!==d.tarih) foreign=true;});
        if(foreign) return false;                             // başka güne ait damga → STALE
        // kesinleşmiş aralık toplamı == kesin (TÜM türler: PDKS, Ücretli, ...)
        const ivs=[];
        document.querySelectorAll('tr').forEach(tr=>{const t=tr.innerText.replace(/\s+/g,' ').trim();
          const m=t.match(/^(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+\d{2}:\d{2}\s+[A-Za-zÇĞİÖŞÜçğıöşü]/);
          if(m)ivs.push([m[1],m[2]]);});
        const toMin=t=>{const mm=/^(\d{1,2}):(\d{2})$/.exec(t);return +mm[1]*60+ +mm[2];};
        const sum=ivs.reduce((s,[a,b])=>s+(toMin(b)-toMin(a)),0);
        return ivs.length>0 && sum===d.inside;
    }"""
    try:
        page.wait_for_function(js, arg={"tarih": tarih, "inside": expected_inside_min}, timeout=8000)
    except PWTimeout:
        raise VerifyError(f"{tarih}: taze dialog gelmedi (STALE korundu — damga tarihi/toplam {to_hhmm(expected_inside_min)} ile eşleşmedi).")
    s = page.evaluate(
        r"""() => {const out=[];
          document.querySelectorAll('tr').forEach(tr=>{const t=tr.innerText.replace(/\s+/g,' ').trim();
            const m=t.match(/^(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+\d{2}:\d{2}\s+[A-Za-zÇĞİÖŞÜçğıöşü]/);
            if(m)out.push([m[1],m[2]]);});
          return JSON.stringify(out);}""")
    rows = json.loads(s)
    return [(to_min(a), to_min(b)) for a, b in rows]


def _set_dialog_time(page: Page, start: int, end: int):
    """Dialog (ekran x<500) Başlangıç/Bitiş select'lerini set et + DOĞRULA."""
    res = page.evaluate(
        r"""({sh,sm,eh,em}) => {
          const setL=(name,v)=>{const el=[...document.querySelectorAll(`select[name="${name}"]`)]
              .filter(e=>e.offsetParent!==null && e.getBoundingClientRect().x<500).pop();
            if(!el) return null;
            el.value=String(v).padStart(2,'0'); el.dispatchEvent(new Event('change',{bubbles:true}));
            if(window.jQuery) jQuery(el).trigger('change'); return el.value;};
          return JSON.stringify({bs:setL('BaslangicSaat',sh), bd:setL('BaslangicDakika',sm),
                                 es:setL('BitisSaat',eh), ed:setL('BitisDakika',em)});
        }""", {"sh": start // 60, "sm": start % 60, "eh": end // 60, "em": end % 60})
    import json
    r = json.loads(res)
    want = {"bs": f"{start//60:02d}", "bd": f"{start%60:02d}", "es": f"{end//60:02d}", "ed": f"{end%60:02d}"}
    if r != want:
        raise VerifyError(f"Dialog saat set edilemedi: oldu={r} beklenen={want}")


def _dialog_open(page: Page) -> bool:
    return page.evaluate(
        r"""() => [...document.querySelectorAll('*')].some(e=>e.offsetParent!==null
              && /Kesinleşmiş Giriş Çıkış Bilgileri/.test(e.textContent) && e.children.length<3)""")


def _close_dialog_hard(page: Page):
    """✎ dialogunu KAYDETMEDEN kapat (skip günleri). SADECE `span.edit-gird-close`'a GERÇEK (Playwright)
    click. KRİTİK BULGU (2026-06-17): Escape veya `.ui-dialog-titlebar .fa-times` FORMU SIFIRLIYOR
    (ilk skip günü tüm temiz tikleri + İstenilen'i 0'lıyordu). Escape ASLA; pencere X'i ASLA.
    Buton tıklamaları flaky → gerçek+JS click dönüşümlü, kapanana kadar poll, form-idle bekle."""
    wait_for_form_idle(page)
    for i in range(6):
        if not _dialog_open(page):
            return
        if i % 2 == 0:
            try:
                page.locator("span.edit-gird-close").filter(visible=True).last.click(timeout=2500)
            except Exception:
                pass
        else:
            page.evaluate(r"""()=>{const c=[...document.querySelectorAll('span.edit-gird-close')].filter(e=>e.offsetParent!==null).pop(); if(c)c.click();}""")
        for _ in range(20):   # ~2 sn tavan kapanma poll'u (100ms granül → kapanmayı erken yakalar)
            page.wait_for_timeout(100)
            if not _dialog_open(page):
                return


def _sweetalert_text(page: Page) -> str:
    """Açık SweetAlert (.sweet-alert) metnini döndür ('' yoksa). Çakışma/hata popup'ı tespiti için."""
    return page.evaluate(
        r"""() => {const s=[...document.querySelectorAll('.sweet-alert')].filter(e=>e.offsetParent!==null)[0];
            return s ? s.innerText.replace(/\s+/g,' ').trim() : '';}""")


def _dismiss_sweetalert(page: Page):
    """Açık SweetAlert'i kapat (TAMAM) + overlay'i temizle — yoksa sonraki işlemleri bloklar."""
    page.evaluate(
        r"""() => {const c=document.querySelector('.sweet-alert button.confirm, .sweet-alert .confirm'); if(c)c.click();
            document.querySelectorAll('.sweet-overlay,.sweet-alert').forEach(e=>e.style.display='none');}""")
    page.wait_for_timeout(400)


def _dialog_save(page: Page):
    """Dialog 'Kaydet' (F3'süz) — gerçek+JS click DÖNÜŞÜMLÜ, kapanana kadar POLL (buton tıklaması flaky).
    Çakışma/hata SweetAlert'i ('pdks kaydı mevcuttur' vb.) çıkarsa KAPAT + net halt (sonraki kişiyi bloklamaz)."""
    wait_for_form_idle(page)
    for i in range(5):
        sa = _sweetalert_text(page)
        if sa:
            _dismiss_sweetalert(page); _close_dialog_hard(page)
            raise VerifyError(f"Dialog Kaydet reddedildi (portal): {sa[:80]}")
        if not _dialog_open(page):
            return
        if i % 2 == 0:
            try:
                page.get_by_text("Kaydet", exact=True).filter(visible=True).last.click(timeout=2500)
            except Exception:
                pass
        else:
            page.evaluate(
                r"""() => {const b=[...document.querySelectorAll('button,a')].filter(x=>x.offsetParent!==null
                      && /Kaydet/i.test(x.textContent) && !/F3/.test(x.textContent)).pop(); if(b)b.click();}""")
        wait_for_form_idle(page)   # Kaydet → AJAX kayıt + grid reload; bitmeden 'kapandı mı' bakma (kullanıcı uyarısı 2026-06-18)
        for _ in range(24):   # ~2.4 sn tavan: kapandı mı / SweetAlert çıktı mı poll (100ms granül → kapanmayı erken yakalar)
            page.wait_for_timeout(100)
            if _sweetalert_text(page) or not _dialog_open(page):
                break
    sa = _sweetalert_text(page)
    if sa:
        _dismiss_sweetalert(page); _close_dialog_hard(page)
        raise VerifyError(f"Dialog Kaydet reddedildi (portal): {sa[:80]}")
    if _dialog_open(page):
        raise VerifyError("Dialog Kaydet tutmadı (dialog hâlâ açık).")


def process_pdks_day(page: Page, tarih: str, target: int) -> str:
    """Bir PDKS gününü işle: satır 'kesin' oku → ✎ aç → STALE-guard oku (çapraz-kontrol) → hesapla →
    set → Kaydet → satırı DOĞRULA → tikle. need<=0 ise dialog kaydetmeden kapatılır, gün tiklenmez."""
    row0 = next((r for r in read_grid_rows(page) if r["tarih"] == tarih), None)
    if not row0 or not row0["kesin"]:
        raise VerifyError(f"{tarih}: PDKS satırı/'kesin' değeri okunamadı.")
    expected_inside = to_min(row0["kesin"])
    # ERKEN SKIP: içeride zaten >= hedef ise (tam izin kesin=09:00, ya da içeride≥9h PDKS) dialog AÇMADAN atla.
    # (Bazı tam-izin günleri disabled gelir / ✎ olmayabilir → gereksiz açmayı önler.)
    if expected_inside >= target:
        log(f"  {tarih}: içeride {row0['kesin']} >= hedef {to_hhmm(target)} → ek yok, atlandı (açılmadı)")
        return "icaride>=hedef"
    if not _open_edit(page, tarih):
        raise VerifyError(f"{tarih}: ✎ ikonu bulunamadı.")
    intervals = _read_dialog(page, tarih, expected_inside)   # çapraz-kontrol: aralık toplamı == kesin
    inside = sum(b - a for a, b in intervals)                # == expected_inside (STALE değil)
    need = target - inside
    if need <= 0:
        # içeride >= hedef → ek yok; dialogu KAYDETMEDEN kapat, gün tiklenmez
        _close_dialog_hard(page)
        if _dialog_open(page):
            raise VerifyError(f"{tarih}: skip dialogu kapatılamadı.")
        log(f"  {tarih}: içeride {to_hhmm(inside)} >= hedef → ek yok, atlandı")
        return "icaride>=hedef"
    # SGK saniye-yuvarlama düzeltmesi: PDKS gününe pdks_buffer dk fazla koy (gün SGK'da 9h'a cap'lenir;
    # form 09:00 ama SGK badge saniyesini aşağı yuvarlayıp 08:59 sayıyordu → +buffer ile net geçer).
    fill = need + CONFIG.get("pdks_buffer", 0)
    # situational başlangıç: kart öğleden sonra başlıyorsa karta bitişik
    first = min(a for a, _ in intervals)
    last = max(b for _, b in intervals)
    prefer = to_hhmm(last + CONFIG["slot_margin"]) if first >= 12 * 60 else CONFIG["prefer_start"]
    start, end = place_outside(fill, intervals, prefer, margin=CONFIG["slot_margin"])
    # çakışma asla olmamalı (savunma)
    for a, b in intervals:
        if not (end <= a or start >= b):
            raise VerifyError(f"{tarih}: pencere {to_hhmm(start)}-{to_hhmm(end)} kartla çakışıyor!")
    _set_dialog_time(page, start, end)
    _dialog_save(page)
    # satır doğrula: toplam == need
    row = next((r for r in read_grid_rows(page) if r["tarih"] == tarih), None)
    if not row:
        raise VerifyError(f"{tarih}: kayıttan sonra satır bulunamadı.")
    if row["toplam"] != to_hhmm(fill):
        raise VerifyError(f"{tarih}: satır toplam {row['toplam']} != gerekli {to_hhmm(fill)}")
    # PDKS günü doldurulduktan SONRA tiklenir
    page.evaluate(
        r"""(d) => {const r=[...document.querySelectorAll('tr')].find(tr=>tr.children[2]&&tr.children[2].textContent.trim()===d);
              const cb=r.querySelector('input[type=checkbox]'); if(cb&&!cb.disabled&&!cb.checked)cb.click();}""", tarih)
    chk = page.evaluate(
        r"""(d) => {const r=[...document.querySelectorAll('tr')].find(tr=>tr.children[2]&&tr.children[2].textContent.trim()===d);
              const cb=r.querySelector('input[type=checkbox]'); return cb?cb.checked:false;}""", tarih)
    if not chk:
        raise VerifyError(f"{tarih}: doldurulduktan sonra tiklenemedi.")
    log(f"  {tarih}: içeride {to_hhmm(inside)} + dışarıda {to_hhmm(fill)} ({to_hhmm(start)}-{to_hhmm(end)}) = {to_hhmm(inside+fill)} (buffer={CONFIG.get('pdks_buffer',0)}dk) ✓")
    return "ok"


def process_arife_day(page: Page, tarih: str, half_target: int):
    """Kurban arifesi (yarım gün resmi tatil). KULLANICI KURALI (2026-06-18): çalışılan yarı = half_target
    (9h kişi → 4:30) DIŞARIDA girilir; diğer yarı otomatik resmi tatil.
      - disabled (badge ile hallolmuş; ALPHAN gibi kesin≈4:30) → ATLA (otomatik).
      - aktif: ✎ aç → (varsa PDKS aralıkları oku; yoksa boş) → need=half_target-içeride → çakışmasız
        yerleştir → kaydet → tikle.
    Dönüş: (durum, eklenen_dışarıda_dk). PDKS gününden farkı: hedef YARIM ve damga OLMAYABİLİR."""
    import json
    row0 = next((r for r in read_grid_rows(page) if r["tarih"] == tarih), None)
    if not row0:
        raise VerifyError(f"{tarih}: arifesi satırı bulunamadı.")
    if row0["disabled"]:
        log(f"  {tarih}: arifesi (badge ile hallolmuş, kesin={row0['kesin'] or '-'}) → atlandı")
        return "atla", 0
    if not _open_edit(page, tarih):
        raise VerifyError(f"{tarih}: arifesi ✎ ikonu bulunamadı.")
    # STALE-guard (ESNEK: damga olmayabilir; olanların tarihi==hedef VE tarih input==hedef olmalı)
    try:
        page.wait_for_function(
            r"""(d) => {
              const xs=[...document.querySelectorAll('input')].filter(i=>i.offsetParent!==null && /^\d{2}\.\d{2}\.\d{4}$/.test((i.value||'').trim()));
              if(xs.length!==1 || xs[0].value.trim()!==d) return false;
              let foreign=false;
              document.querySelectorAll('tr').forEach(tr=>{const t=tr.innerText.replace(/\s+/g,' ').trim();
                const m=t.match(/^(\d{2}\.\d{2}\.\d{4})\s+\d{2}:\d{2}\s+\S+\s+(PDKS|Manuel|Elle)/);
                if(m && m[1]!==d) foreign=true;});
              return !foreign;
            }""", arg=tarih, timeout=8000)
    except PWTimeout:
        raise VerifyError(f"{tarih}: arifesi dialog'u taze gelmedi (STALE).")
    page.wait_for_timeout(300)
    s = page.evaluate(
        r"""() => {const out=[];document.querySelectorAll('tr').forEach(tr=>{const t=tr.innerText.replace(/\s+/g,' ').trim();
          const m=t.match(/^(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+(PDKS|Manuel|Elle)/); if(m)out.push([m[1],m[2]]);});return JSON.stringify(out);}""")
    intervals = [(to_min(a), to_min(b)) for a, b in json.loads(s)]
    inside_total = sum(b - a for a, b in intervals)
    # 🔴 ARİFE (CANLI BULGU 2026-06-18 ŞAHİN SAMET EKİN, SGK raporu): içeride PDKS — sabah work-half'ta olsa
    # BİLE — istisnaya SAYILMIYOR; rapor YALNIZ beyan edilen DIŞARIDA'yı sayar (ŞAHİN: içeride 2:40, beyan 1:50
    # → rapor 1:50 → Gün 28,70). Bu yüzden TAM half_target (4:30) DIŞARIDA beyan edilir (içeriyle çakışmamak
    # için place_outside busy=içeride). CANER (içeride 0 → 4:30 beyan) zaten Gün=30 almıştı. İçeride = bilgisel.
    need = half_target
    start, end = place_outside(need, intervals, CONFIG["prefer_start"], margin=CONFIG["slot_margin"])
    for a, b in intervals:
        if not (end <= a or start >= b):
            raise VerifyError(f"{tarih}: arifesi penceresi kartla çakışıyor!")
    _set_dialog_time(page, start, end)
    _dialog_save(page)
    row = next((r for r in read_grid_rows(page) if r["tarih"] == tarih), None)
    if not row or row["toplam"] != to_hhmm(need):
        raise VerifyError(f"{tarih}: arifesi satır toplam {row['toplam'] if row else '?'} != {to_hhmm(need)}")
    page.evaluate(
        r"""(d) => {const r=[...document.querySelectorAll('tr')].find(tr=>tr.children[2]&&tr.children[2].textContent.trim()===d);
              const cb=r.querySelector('input[type=checkbox]'); if(cb&&!cb.disabled&&!cb.checked)cb.click();}""", tarih)
    chk = page.evaluate(
        r"""(d) => {const r=[...document.querySelectorAll('tr')].find(tr=>tr.children[2]&&tr.children[2].textContent.trim()===d);
              const cb=r.querySelector('input[type=checkbox]'); return cb?cb.checked:false;}""", tarih)
    if not chk:
        raise VerifyError(f"{tarih}: arifesi doldurulduktan sonra tiklenemedi.")
    log(f"  {tarih}: arifesi (içeride sayılmaz: {to_hhmm(inside_total)}) → dışarıda {to_hhmm(need)} ({to_hhmm(start)}-{to_hhmm(end)}) = {to_hhmm(half_target)} ✓")
    return "ok", need


def reopen_fresh_form(page: Page):
    """Sonraki kişi için TEMİZ kart (canlı 2026-06-17 doğrulandı):
      1) AÇIK FORM penceresini kapat (jQuery UI .fa-times).
      2) Grid'in DOM'dan gittiğini teyit et — gitmezse iki grid olur, querySelectorAll iki kişinin
         satırlarını karıştırır (felaket). Kapanmadıysa menüden temiz aç (yedek).
      3) Arkadaki LİSTE penceresini yeniden kullan: Dönem=151 + liste 'Yeni (F2)'.
    Form toolbar'ının 'Yeni (F2)'sini KULLANMA (autocomplete bozulur). Sayfa RELOAD YOK (Cloudflare)."""
    # 1) açık form penceresini kapat
    page.evaluate(
        r"""() => {const bar=[...document.querySelectorAll('div.ui-dialog-titlebar')]
              .filter(b=>b.offsetParent!==null && /Dışarıda Geçirilen Süreler Formu/.test(b.textContent))[0];
            if(bar){const x=bar.querySelector('.fa-times, .ui-dialog-titlebar-close'); if(x)x.click();}}""")
    page.wait_for_timeout(700)
    # 2) grid gerçekten kalktı mı?
    grid_left = page.evaluate(
        r"""(re)=>[...document.querySelectorAll('tr')].filter(tr=>tr.children[2]&&new RegExp(re).test((tr.children[2].textContent||'').trim())).length""",
        CONFIG["ay_regex"])
    has_list = page.evaluate(
        r"""()=>[...document.querySelectorAll('#Donem_Id, select[name="Donem_Id"]')].some(e=>e.offsetParent!==null)""")
    if grid_left > 0 or not has_list:
        # form kapanmadı / liste yok → menüden temiz aç (yedek yol)
        open_dgs_form(page)
    # 3) listeyi yeniden kullan
    list_set_donem_and_yeni(page)
    ensure_card_fresh(page)


def close_nav_menu(page: Page):
    """Açık navbar (PERSONEL vb.) dropdown'ını KAPAT.
    🔴 CANLI BULGU (2026-07-14, TPI — ARDA DENİZ BOSTAN, kullanıcı yakaladı): `_open_menu_item` menüyü
    tıkladıktan sonra dropdown AÇIK kalabiliyor. Menü kabı `div#caixa` üst şeridi kaplıyor (ölçüm: y=86..463,
    tam genişlik) ve içindeki <a> menü öğeleri form toolbar'ının (Kaydet / Onaya Gönder) ÜSTÜNE biniyor →
    GERÇEK Playwright click "<a> from <div id='caixa'> subtree intercepts pointer events" ile 5sn TIMEOUT yer.
    Bu portalda 'Onaya Gönder' GERÇEK click İSTER (evaluate-click jQuery handler'ını tetiklemez) → menü açıksa
    onay GİTMEZ, taslak kalır.
    ⚠️ Escape KULLANMA — jQuery ui-dialog'u (açık formu) kapatır.
    ⚠️ Sentetik body-click / mouseleave İŞE YARAMIYOR (canlı denendi, menü açık kaldı) — menü JS click-toggle.
    ÇALIŞAN YÖNTEM (canlı ölçüm): menü kütüphanesi KAPALI alt-menüleri kendi inline `style="display:none"`ı ile
    tutuyor (tüm diğer `[id^=menu-folder]`'lar öyle) → açık kalanı AYNI gösterimle kapat. Sonraki açılışta
    kütüphanenin `.show()`/display:block'u bu inline stili ezer (menü bozulmaz). Ayrıca gerçek imleci menüden
    uzaklaştır (olası hover state)."""
    try:
        page.mouse.move(5, 720)
    except Exception:
        pass
    page.evaluate(
        r"""() => {
          document.querySelectorAll('[id^=menu-folder]').forEach(f => {
            if (getComputedStyle(f).display !== 'none') f.style.display = 'none';
          });
        }""")
    page.wait_for_timeout(150)


def _visible_form_dialogs(page: Page) -> int:
    """Görünür 'Dışarıda Geçirilen Süreler Formu' dialog SAYISI (stale/çift form tespiti)."""
    return page.evaluate(
        r"""()=>[...document.querySelectorAll('div.ui-dialog')].filter(d=>d.offsetParent!==null
              && /Dışarıda Geçirilen Süreler Formu/.test(d.querySelector('.ui-dialog-titlebar')?.textContent||'')).length""")


def save_draft(page: Page) -> str:
    """Form 'Kaydet (F3)' → TASLAK. 'başarı ile kayıt' mesajını DOĞRULA. dry_run'da tıklamaz.
    🔴 CANLI BULGU (2026-07-14, TPI — MÜKERRER KAYIT SEBEBİ): eski hâli Kaydet'ten sonra SABİT 1.5sn bekleyip
    mesajı TEK KEZ kontrol ediyordu. Sayfa yüklüyken (birden çok dialog açık) başarı mesajı 1.5sn'den GEÇ geldi
    → "kaydedilmedi" HATASI verdi ama kayıt SERVER'A GİTMİŞTİ (yanlış-negatif) → operatör tekrar denedi →
    ARDA DENİZ BOSTAN için İKİ mükerrer taslak (610360+610361). DERS (Yıldız/GÜLŞAH ile aynı): 'mesaj görülmedi'
    ≠ 'kaydolmadı'. ÇÖZÜM: mesajı ~10sn POLL et + Kaydet'i AÇIK FORMUN İÇİNDEN tıkla (global ilk eşleşme değil —
    iki form açıksa yanlış forma basıyordu) + çift form varsa DUR (verify-or-halt)."""
    if CONFIG["dry_run"]:
        return "DRY-RUN (kaydedilmedi)"
    close_nav_menu(page)   # açık dropdown Kaydet/Onaya Gönder'in üstüne biniyor (2026-07-14)
    n_dlg = _visible_form_dialogs(page)
    if n_dlg != 1:
        raise VerifyError(f"Kaydet öncesi {n_dlg} adet DGS formu açık (1 olmalı) — stale dialog, "
                          f"yanlış forma kaydetme riski. MANUEL kontrol.")
    clicked = page.evaluate(
        r"""() => {const dlg=[...document.querySelectorAll('div.ui-dialog')].filter(d=>d.offsetParent!==null
              && /Dışarıda Geçirilen Süreler Formu/.test(d.querySelector('.ui-dialog-titlebar')?.textContent||''))[0];
            if(!dlg) return false;
            const b=[...dlg.querySelectorAll('a,button')].find(x=>x.offsetParent!==null
              && /Kaydet/i.test(x.textContent) && /F3/.test(x.textContent));
            if(!b) return false; b.click(); return true;}""")
    if not clicked:
        raise VerifyError("Açık formda 'Kaydet (F3)' butonu bulunamadı.")
    # başarı mesajını POLL et (sabit bekleme YOK — yüklü sayfada mesaj geç gelebiliyor → yanlış-negatif → MÜKERRER)
    for _ in range(20):
        page.wait_for_timeout(500)
        if "başarı ile kayıt" in page.inner_text("body").lower():
            return "TASLAK kaydedildi"
    raise VerifyError("Form Kaydet sonrası 'başarı ile kayıt' mesajı 10sn içinde görülemedi. "
                      "⚠️ KAYIT GİTMİŞ OLABİLİR — yeniden girmeden ÖNCE listeyi kontrol et (mükerrer riski).")


def form_status(page) -> str:
    """AÇIK formdaki 'Onay Durumu' select metni (onay doğrulaması için anchor)."""
    return page.evaluate(
        r"""()=>{const dlg=[...document.querySelectorAll('div.ui-dialog')].filter(d=>d.offsetParent!==null && /Dışarıda Geçirilen Süreler Formu/.test(d.querySelector('.ui-dialog-titlebar')?.textContent||''))[0];
            if(!dlg)return ''; const sels=[...dlg.querySelectorAll('select')].map(s=>(s.selectedOptions[0]?.text||'').trim())
              .filter(t=>/Gönderil|Onay|Beklen|Reddedil|Değerlendir/.test(t)); return sels.join(' | ');}""")


def onaya_gonder_in_form(page) -> str:
    """Kaydet'ten SONRA, AÇIK formdaki 'Onaya Gönder' butonuna bas (buton Kaydet'in DİBİNDE) + onay istemini
    geç + statü taslaktan ('Gönderilmemiş') çıktı mı DOĞRULA. (dgs_onaya.click_onaya_gonder ile aynı mantık;
    form zaten bu kişinin → save_draft kimliği doğrulanmış kişiyi kaydetti, ek personel kontrolü gerekmez.)
    KULLANICI İSTEĞİ (2026-06-19): 'kaydetip onaya yollamıyordu; onay butonu dibinde, kabulden ~0.5sn sonra bas.'
    Onay GİDEMEZSE VerifyError → çağıran taslağı korur (dgs_onaya ile sonradan gönderilebilir)."""
    state = page.evaluate(
        r"""()=>{const b=[...document.querySelectorAll('a,button')].find(x=>x.offsetParent!==null && /Onaya Gönder/i.test(x.textContent));
            if(!b)return 'yok';
            const dis=b.disabled || /disabled/.test(b.className) || b.getAttribute('aria-disabled')==='true';
            return dis?'pasif':'aktif';}""")
    if state != "aktif":
        raise VerifyError(f"'Onaya Gönder' {state} (kaydet sonrası buton hazır değil).")
    # GERÇEK Playwright click (evaluate-click jQuery handler'ı GÜVENİLMEZ tetikler — dgs_onaya bulgusu).
    # Açık navbar dropdown'ı bu butonun ÜSTÜNE biniyor ve click'i yiyor ("#caixa subtree intercepts pointer
    # events", 2026-07-14 TPI) → önce menüyü kapat; yine engellenirse bir kez daha kapatıp tekrar dene.
    btn = page.locator("a:has-text('Onaya Gönder'), button:has-text('Onaya Gönder')").filter(visible=True).first
    close_nav_menu(page)
    try:
        btn.click(timeout=5000)
    except Exception as e:
        if "intercepts pointer events" not in str(e):
            raise
        log("  (Onaya Gönder tıklaması engellendi → navbar menüsü kapatılıp tekrar deneniyor)")
        close_nav_menu(page)
        btn.click(timeout=5000)
    wait_for_form_idle(page)
    # onay istemi PORTAL-BAĞIMLI: Bootstrap (#modalYes/.color-green) / SweetAlert / native confirm (dialog handler accept eder)
    for sel in ("#modalYes", ".modal.in button.color-green", ".modal.show button.color-green",
                ".modal.in button:has-text('Evet')", ".sweet-alert button.confirm", ".sweet-alert a.confirm"):
        try:
            page.locator(sel).filter(visible=True).first.click(timeout=2000)
            break
        except Exception:
            continue
    else:
        for nm in ("Evet", "Eminim", "Onayla"):
            try:
                page.get_by_role("button", name=nm, exact=True).filter(visible=True).first.click(timeout=1200)
                break
            except Exception:
                continue
    wait_for_form_idle(page)
    # DOĞRULA: statü taslaktan çıkana dek ~6sn poll (async güncelleme → false-negative önle)
    st = ""
    for _ in range(6):
        st = form_status(page)
        if st and "Gönderilmemiş" not in st:
            return st
        page.wait_for_timeout(1000)
    sa = _sweetalert_text(page)
    if sa:   # 'dilekçe zorunlu' / çakışma vb. → taslak duruyor, halt
        _dismiss_sweetalert(page)
        raise VerifyError(f"Onaya Gönder reddedildi (portal): {sa[:80]}")
    raise VerifyError(f"Onaya gitmedi (statü taslakta kaldı: {st!r}).")


def process_one_person(page: Page, person: PersonRow, first: bool, sched_map: dict) -> dict:
    """Tek kişiyi uçtan uca. TÜM kontroller geçmeden Kaydet YOK. Hata → kaydedilmez, FAILED döner."""
    if not is_ar_ge(person) and not CONFIG.get("include_destek"):
        return {"ad": person.ad_soyad, "ok": True, "mesaj": f"ATLANDI (tür={person.ar_ge}, Ar-Ge değil)"}

    # kişi-bazında haftalık program → günlük hedef + cumartesi. Standart dışıysa BURADA DURUR (yanlış girmez).
    daily_target, include_sat = person_schedule(person.ad_soyad, sched_map)

    if first:
        open_dgs_form(page)
        list_set_donem_and_yeni(page)
        ensure_card_fresh(page)
    else:
        reopen_fresh_form(page)

    ensure_gorevlendirme_diger(page)
    select_personel(page, person.ad_soyad, person.tc)
    ensure_calisma_turu(page)
    set_proje(page, person.proje_adi)   # Proje_Id doğrulanır
    if include_sat:
        enable_saturday(page)           # 6 günlük: cumartesileri grid'e ekle (allow_6day gerektirir)

    rows = read_grid_rows(page)
    # SINIFLANDIRMA AÇIKLAMA (td[12]) İLE — kesin İLE DEĞİL (CANLI BULGU 2026-06-18, izin girildikten sonra):
    #   "Pdks Kaydı mevcuttur"  → PDKS (✎ ile işle)
    #   "İzin Kaydı mevcuttur"  → İZİN (atla; portal kesin=09:00 ile otomatik kredi veriyor)
    #   "...resmi tatili/Bayram" → TATİL (atla; yarım-tatil arifesi kesin=04:30 gelebilir ama yine atla)
    #   açıklama boş + aktif + kesin yok → TEMİZ (tikle)
    # (Eski 'pdks = kesin var' yanlıştı: izin=09:00 ve yarım-tatil=04:30 yanlışça PDKS sanılıyordu.)
    # KRİTİK (2026-06-18): açıklama (td[12]) bazen BOŞ gelir PDKS gününde bile (örn. BARIŞ ÖZÇİMEN 12.05,
    # kesin=01:09 ama açıklama boş). Bu yüzden PDKS'i: açıklamada "pdks" VEYA (kesin var VE izin/arife/tatil
    # DEĞİL) diye sınıfla. Yoksa disabled-boş-açıklama-PDKS günü "skip"e düşer → hesap+İstenilen ikisi de
    # dışlar → eşleşir ve EKSİK kaydeder (sessiz hata!). İzin/arife/tatil açıklamaları önce elenir.
    # NOT (Türkçe İ): "İzin".lower() == "i̇zin" (i + U+0307 birleşik nokta) → "izin" ile EŞLEŞMEZ!
    # U+0307'yi temizleyerek normalize et (yoksa is_izin tutmaz → izin günü PDKS sanılır → yarım izinde
    # 4:30 fazla girilir). CANLI BULGU 2026-06-18 (BERKE 04.05 yarım izin).
    a = lambda r: (r["aciklama"] or "").lower().replace("̇", "")
    is_arife = lambda r: "arife" in a(r)                                  # Kurban arifesi → yarım gün
    is_izin = lambda r: "izin" in a(r)                                    # izin → atla (oto kredi)
    is_tatil = lambda r: (("tatil" in a(r)) or ("bayram" in a(r))) and not is_arife(r)
    # İZİN günü de "doldurulacak" sınıfa girer (kullanıcı kuralı 2026-06-18: yarım izinde çalışma yarısı
    # girilir; tam izinde need=0 → atlanır). İzin, kesinleşmiş tabloda "Ücretli" aralığı olarak görünür →
    # process_pdks_day need = 9h − içeride(Ücretli+PDKS) hesaplar. tatil/arife hariç tutulur.
    is_pdks = lambda r: (not is_arife(r) and not is_tatil(r)
                         and ("pdks" in a(r) or "izin" in a(r) or bool(r["kesin"])))
    pdks = [r for r in rows if is_pdks(r)]
    arife = [r for r in rows if is_arife(r)]
    temiz = [r for r in rows if not is_pdks(r) and not is_arife(r) and not is_izin(r) and not is_tatil(r)
             and not r["disabled"] and not r["kesin"] and not r["aciklama"]]
    handled = {r["tarih"] for r in pdks} | {r["tarih"] for r in arife} | {r["tarih"] for r in temiz}
    skip = [r for r in rows if r["tarih"] not in handled]   # izin / tam tatil → dokunma
    # GÜVENLİK: atlanacaklar arasında AKTİF (disabled olmayan) bir gün varsa bu bilinmeyen bir durum → DUR
    weird = [r for r in skip if not r["disabled"]]
    if weird:
        raise VerifyError(f"{person.ad_soyad}: beklenmedik aktif gün(ler) "
                          f"{[r['tarih']+'('+(r['aciklama'] or 'boş')+')' for r in weird]} — MANUEL incele.")
    log(f"  Grid: {len(rows)} gün | temiz={len(temiz)} pdks={len(pdks)} arife={len(arife)} "
        f"atla(izin/tatil)={len(skip)} | hedef={to_hhmm(daily_target)}/gün cmt={include_sat}")

    # ÇAPRAZ KONTROL: temiz gün varsayılan toplamı, programdan türetilen günlük hedefle uyuşmalı.
    # (Uymazsa program/grid çelişkisi var → yanlış girmektense DUR.)
    defs = {r["toplam"] for r in temiz if r["toplam"]}
    if defs and defs != {to_hhmm(daily_target)}:
        raise VerifyError(f"{person.ad_soyad}: temiz gün varsayılanı {sorted(defs)} != program hedefi "
                          f"{to_hhmm(daily_target)} — program/grid uyuşmazlığı, MANUEL.")

    # temiz günler (proje seçili olduğu için tutar)
    n_clean = tick_clean_days(page)
    expected = n_clean * daily_target

    # PDKS günleri (kişi-bazında hedefle)
    for r in pdks:
        res = process_pdks_day(page, r["tarih"], daily_target)
        if res == "ok":
            row = next((x for x in read_grid_rows(page) if x["tarih"] == r["tarih"]), None)
            if row and row["toplam"]:
                expected += to_min(row["toplam"])

    # ARİFESİ günleri (yarım gün = daily_target/2; kullanıcı kuralı 2026-06-18)
    half_target = daily_target // 2
    for r in arife:
        _st, added = process_arife_day(page, r["tarih"], half_target)
        expected += added

    # SON KONTROL: tikli gün sayısı + İstenilen toplam beklenenle uyuyor mu?
    after = read_grid_rows(page)
    ticked = [r for r in after if r["checked"]]
    istenilen = page.evaluate(
        r"""() => {const c=document.querySelector('.BuFordaIstenilen'); const i=c?c.querySelector('input'):null; return i?i.value.trim():'?';}""")
    if istenilen != "?" and to_min(istenilen) != expected:
        raise VerifyError(f"İstenilen toplam uyuşmuyor: portal={istenilen} hesap={to_hhmm(expected)}")
    log(f"  KONTROL: tikli={len(ticked)} | İstenilen={istenilen} (beklenen {to_hhmm(expected)}) ✓")

    msg = save_draft(page)
    onaylandi = False
    if not CONFIG["dry_run"] and CONFIG.get("auto_onay"):
        page.wait_for_timeout(500)   # kullanıcı kuralı: kabul/kaydet otursun, ~0.5sn SONRA onaya bas
        wait_for_form_idle(page)     # kaydet AJAX'ı tam bitsin (buton aktifleşsin)
        try:
            st = onaya_gonder_in_form(page)
            msg = f"{msg} + ONAYA GÖNDERİLDİ ({st})"
            onaylandi = True
        except VerifyError as e:
            # onay gidemedi → taslak DURUYOR (kaydedildi); dgs_onaya ile sonradan gönderilebilir. Kişi atlanmaz.
            msg = f"{msg} | ⚠ ONAYA GÖNDERİLEMEDİ: {e} (taslak DURUYOR → dgs_onaya ile gönder)"
    return {"ad": person.ad_soyad, "ok": True, "mesaj": msg, "istenilen": istenilen, "onaylandi": onaylandi}


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="DGS Otomasyon (verify-or-halt; TASLAK; Onaya Gönder insanda)")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--sheet", default=None,
                    help="Excel sayfa adı. Verilmezse dönemin Türkçe ay adına göre türetilir (ör. 'Haziran').")
    ap.add_argument("--donem", default=None,
                    help="Dönem etiketi override (ör. 'HAZİRAN 2026'); yoksa bugünden bir önceki ay "
                         "(dgs_onaya.py/dgs_rapor_kontrol.py ile aynı otomatik desen).")
    ap.add_argument("--lokasyon", default="TPI", help="Lokasyonu SGK filtresi (TPI, ARI, Ulutek...)")
    ap.add_argument("--person", default=None, help="Sadece tek kişi (test)")
    ap.add_argument("--file", default=None, help="Belirli isimler (her satır bir ad); done-skip atlanır")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--commit", action="store_true", help="TASLAK kaydet (Onaya Gönder YİNE YOK)")
    ap.add_argument("--dump-names", default=None, help="DEBUG: program haritası isimlerini (alt-dize filtreli) bas ve çık")
    ap.add_argument("--assume-std", action="store_true", help="Haritada olmayan kişide standart 5g/45s (540) varsay (grid ile doğrulanır)")
    ap.add_argument("--include-destek", action="store_true", help="Ar-Ge olmayan (Destek) kişileri de işle")
    ap.add_argument("--onayla", action="store_true",
                    help="Kaydet'ten ~0.5sn sonra formdaki 'Onaya Gönder'e bas (taslak+onay TEK geçişte; buton "
                         "Kaydet'in dibinde). Onay gidemezse taslak DURUR (dgs_onaya ile gönderilebilir). "
                         "Başarılı onaylar dgs_onaya_done_<lok>_<sheet>.txt'e de yazılır (dgs_onaya çift-göndermez).")
    ap.add_argument("--no-schedule", action="store_true",
                    help="read_schedule_map'i ATLA (40-sayfa Personel Listesi okuması; çoğu parkta harita boş "
                         "geliyor zaten). Herkes assume-std 540; grid temiz-gün çapraz-kontrolü (==09:00) "
                         "6-günlüğü/anomaliyi yakalar. --assume-std'yi otomatik açar. İlk kişide --limit 1 ile doğrula.")
    args = ap.parse_args()
    CONFIG["dry_run"] = not args.commit
    CONFIG["assume_standard"] = args.assume_std or args.no_schedule
    CONFIG["include_destek"] = args.include_destek
    CONFIG["auto_onay"] = args.onayla

    donem_label, ay_regex, ay_no = donem_label_and_regex(args.donem)
    CONFIG["donem_text"] = donem_label
    CONFIG["ay_regex"] = ay_regex
    sheet = args.sheet or TR_AYLAR_TITLE[ay_no]
    log(f"Dönem: {donem_label} (ay_regex={ay_regex}, sheet='{sheet}')")

    allp = read_excel(args.excel, sheet)
    if args.person:
        key = args.person.upper()
        if key not in allp:
            log(f"HATA: '{args.person}' yok. Örnek: {list(allp)[:5]}"); sys.exit(1)
        targets = [allp[key]]
    elif args.file:
        names = [l.strip() for l in open(args.file, encoding="utf-8") if l.strip()]
        targets = []
        for nm in names:
            k = nm.upper()
            if k in allp:
                targets.append(allp[k])
            else:
                log(f"  UYARI: '{nm}' Excel'de yok, atlandı.")
    else:
        targets = [p for p in allp.values() if p.lokasyon.upper() == args.lokasyon.upper() and is_ar_ge(p)]
    bypass_done = bool(args.person or args.file)   # --person/--file: done-skip ATLA (yeniden-girme için)
    if args.limit:
        targets = targets[: args.limit]
    log(f"Hedef: {len(targets)} kişi (lokasyon={args.lokasyon}, Ar-Ge) | commit={args.commit} (False=DRY-RUN)")
    if not targets:
        sys.exit(1)

    done_file = f"dgs_done_{args.lokasyon}_{sheet}.txt"
    done = set(open(done_file, encoding="utf-8").read().splitlines()) if os.path.exists(done_file) else set()
    if done:
        log(f"{len(done)} kişi zaten işlenmiş (atlanacak).")

    results = []
    with sync_playwright() as pw:
        browser, page = attach_browser(pw)
        if CONFIG.get("auto_onay"):
            page.on("dialog", lambda d: d.accept())   # 'Onaya Gönder' native confirm() çıkarsa kabul (submit iptal etme)
        try:
            assert_logged_in(page)
            # Haftalık program haritasını BİR KEZ oku (5 gün/45 saat mi, 6 gün mü) — kişi-bazında hedef için.
            # --no-schedule: ATLA (40-sayfa okuma; harita çoğu parkta boş) → assume-std + grid çapraz-kontrol.
            if args.no_schedule:
                sched_map = {}
                log("schedule_map ATLANDI (--no-schedule): herkes assume-std 540; grid temiz-gün==09:00 "
                    "çapraz-kontrolü her kişide 6-günlüğü/anomaliyi yakalar (uymazsa DURUR).")
            else:
                sched_map = read_schedule_map(page)
            if args.dump_names is not None:
                sub = args.dump_names.upper()
                hits = sorted(k for k in sched_map if sub in k)
                log(f"DUMP '{args.dump_names}' → {len(hits)} eşleşme:")
                for k in hits:
                    log(f"    [{k}] = {sched_map[k]}")
                return
            first = True
            for i, person in enumerate(targets, 1):
                if person.ad_soyad in done and not bypass_done:
                    continue
                log(f"\n=== [{i}/{len(targets)}] {person.ad_soyad} (proje: {person.proje_adi[:35]}) ===")
                try:
                    assert_logged_in(page)  # her kişiden önce oturum teyidi
                    r = process_one_person(page, person, first, sched_map)
                    first = False
                except CloudflareHalt as e:
                    log(f"!! DURDU (Cloudflare/oturum): {e}")
                    log("   Elle Cloudflare'i geç + dashboard'a dön, sonra tekrar koş (resume devam eder).")
                    break
                except VerifyError as e:
                    ts = int(time.time()); shot = f"dgs_verify_{ts}.png"
                    try: page.screenshot(path=shot, full_page=True)
                    except Exception: pass
                    r = {"ad": person.ad_soyad, "ok": False, "mesaj": f"DOĞRULAMA HATASI (KAYDEDİLMEDİ): {e} (ss:{shot})"}
                    # kart yarım kaldı; sonraki kişi reopen_fresh_form ile temiz başlar
                    first = False
                except Exception as e:
                    ts = int(time.time()); shot = f"dgs_hata_{ts}.png"
                    try: page.screenshot(path=shot, full_page=True)
                    except Exception: pass
                    r = {"ad": person.ad_soyad, "ok": False, "mesaj": f"HATA (KAYDEDİLMEDİ): {e} (ss:{shot})"}
                    first = False
                results.append(r)
                log(f"  -> {r['mesaj']}")
                if r["ok"] and args.commit and "ATLANDI" not in r["mesaj"] and "DRY-RUN" not in r["mesaj"]:
                    with open(done_file, "a", encoding="utf-8") as f:
                        f.write(person.ad_soyad + "\n")
                    if r.get("onaylandi"):   # --onayla başarılı → onay-resume'a da yaz (dgs_onaya çift-göndermez)
                        with open(f"dgs_onaya_done_{args.lokasyon}_{sheet}.txt", "a", encoding="utf-8") as f:
                            f.write(person.ad_soyad + "\n")

            ok = sum(1 for r in results if r["ok"])
            log(f"\n==== ÖZET: {ok}/{len(results)} sorunsuz ====")
            for r in results:
                if not r["ok"]:
                    log(f"  !! BAŞARISIZ: {r['ad']} — {r['mesaj']}")
            if CONFIG.get("auto_onay"):
                onaylanan = sum(1 for r in results if r.get("onaylandi"))
                log(f"HATIRLATMA: --onayla açıktı → {onaylanan}/{len(results)} kişi TASLAK+ONAYA GÖNDERİLDİ. "
                    f"Onaya gidemeyen taslaklar için: dgs_onaya.py --lokasyon {args.lokasyon} --commit")
            else:
                log("HATIRLATMA: Hepsi TASLAK. 'Onaya Gönder' ayrı: dgs_onaya.py (veya --onayla ile tek geçişte).")
            log(f"DOĞRULAMA: PERSONEL > PDKS > 'SGK Çalışan Bildirgesi Gün Detaylı Raporu' ({donem_label}) → herkeste 'Gelir Vergisi İstisnası Gün' = 30 olmalı.")
            try:
                input(f"{LOG} >> Bitti. ENTER ile kapat...")
            except EOFError:
                pass
        except CloudflareHalt as e:
            log(f"GENEL DURUŞ (Cloudflare/oturum): {e}")
        except Exception as e:
            log(f"GENEL HATA: {e}")
            raise


if __name__ == "__main__":
    main()
